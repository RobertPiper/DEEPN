#!/usr/bin/env python3
import os
import sys

if '.app/Contents/MacOS' in sys.executable:
    os.chdir(os.path.join(os.path.dirname(sys.executable), '..', 'Resources'))

import re
import csv
import math
import time
import glob
import tempfile
from datetime import datetime

from PyQt5 import QtCore, QtGui, QtWidgets, uic
import xlsxwriter as xls
import openpyxl

import functions.fileio_gui as f
import functions.stat_collation as sc
import DragDropListView

app = QtWidgets.QApplication(sys.argv)
form_class, base_class = uic.loadUiType(os.path.join('ui', 'Stat_Maker_v3.ui'))

R_SCRIPT_PATH = os.path.join('r_scripts', 'run_Y2H_enrichement_stats_v5.R')
BAYESIAN_R_SCRIPT_PATH = os.path.join('r_scripts', 'run_Y2H_bayesian_stats_v3.R')
BAYESIAN_THRESHOLD_PPM = 3  # matches Stat Maker v1/v2's threshold_sbx default

METHOD_INFO_TEXT = """STAT MAKER v7 - METHOD OVERVIEW

WHAT THIS PROGRAM DOES

Stat Maker compares one or two Y2H baits (each Selected vs Non-Selected)
against a Vector control (2 replicates, Selected vs Non-Selected each),
then combines the results into one table, one row per gene:

  1. STATISTICS (via DESeq2, see R code below)
     For each gene: pvalue, log2FoldChange, and a rank-based Enrichment_score,
     computed separately for Bait1 (Selected vs Non-Selected), Vector
     (Selected vs Non-Selected), and - if a Bait2 dataset is also
     supplied - for Bait2 as well, plus a Bait1-vs-Bait2 specificity
     contrast (pvalue_specificity/log2FoldChange_specificity: positive
     means more enriched under Bait1 selection than Bait2 selection, i.e.
     more specific to Bait1; negative means the opposite).

     Bait2 is optional. With just Bait1 + Vector supplied, this produces
     the exact same Bait1/Vector columns Stat Maker v2 always did (just
     named "_bait1" instead of "_bait") - no specificity columns, since
     there's no second bait to compare against.

  2. JUNCTION COLLATION (from BLAST results, .bqp files)
     For each gene, in each input dataset, the percentage of junctions
     that are:
       inframe_inorf          - in-frame AND within the annotated ORF
       upstream               - start upstream of the ORF
       in_orf                 - within the ORF (regardless of frame)
       downstream             - start downstream of the ORF
       in-frame:forward       - in-frame, forward orientation (regardless of ORF position)
       out-of-frame:forward   - out-of-frame, forward orientation (regardless of ORF position)
       backwards              - mapped in the reverse orientation
       intron                 - ORF contains an intron (yeast genes only; rare)

     Frame and orientation are not independently tracked: the underlying
     classification (functions/junctionf_gui.py) computes in-frame/
     out-of-frame first, then overwrites that with "backwards" if the
     junction is reverse-oriented, or "intron" if the gene contains one -
     so "backwards" and "intron" junctions carry no frame information (a
     backwards junction could have been in-frame or not; that's discarded).
     "in-frame:forward"/"out-of-frame:forward" are named accordingly rather
     than "in_frame"/"not_in_frame", since forward orientation is implied,
     not just frame status - a junction can be positionally in-frame and
     backwards at the same time, and that case is reported as "backwards"
     only, not as "in-frame:forward".

INPUT

  Gene-level counts come from DEEPN's gene_count_summary CSVs, which report
  PPM (parts per million of total mapped reads), not raw read counts. Genes
  annotated on multiple scaffolds/alt-contigs are summed to one PPM total
  per gene before anything else happens.

OUTPUT TABLE LAYOUT

  Above the Gene table: each input dataset's path, then OVERDISPERSION
  (both methods' estimates, see below), then CRUSH (see next section).

  The Gene table itself has three header rows - a block-title row, a
  per-column scale/unit row (Linear FC / log2 / P / p-value / Percent),
  and the field-name row the sort/filter dropdowns attach to - followed by
  three blocks of per-gene columns, each block color-coded by which method
  a column comes from (blue = DESeq2, purple = Bayesian/MCMC, pink = the
  Select junction-% column):

    Bait1 stats / Bait2 stats (Bait2 only if a second bait was supplied):
      ppm non, ppm select                    - that bait's own PPM
      raw enrichment (ppm select/non)         - linear fold change, no model
      Enr1/Enr2                               - raw log2 fold change (model-free,
                                                 from summary.data.deepn())
      AdjEnr1/AdjEnr2                         - Bayesian/MCMC posterior median
                                                 log2 effect size (Bait vs Vector)
      DESeq2 (BaitN vs Vector) [log2 FC row]  - DESeq2's log2FoldChange for the
                                                 same Bait-vs-Vector contrast
      pBaitN_Vec                              - Bayesian posterior probability
      DESeq2 (BaitN vs Vector) [p-value_raw]  - DESeq2's p-value, raw counts +
                                                 poscounts normalization
      DESeq2 (BaitN vs Vector) [p-value_norm] - DESeq2's p-value, normalized
                                                 PPM data (Y2H-SCORES method)
      in-frame:Forward (Select/Non)           - % of that dataset's junctions
                                                 in-frame, forward orientation

    Vector stats:
      Same ppm non/select/raw-enrichment/DESeq2/junction columns as above,
      but every value is the mean of Vector1 and Vector2 (the two vector
      replicates) - ppm non/ppm select are each replicate's PPM averaged
      per gene first, then the fold change is computed from those two
      averaged PPMs (not averaged as two separate fold-change ratios).
      Two extra Non-Select-only columns (out-frame:Forward, backward)
      characterize the starting prey pool's own junction quality, shared
      across Bait1/Bait2's non-selected pools too since it's the same
      library before any selection pressure is applied.

  To the right of that: the original, more granular report this table
  used to be entirely - DESeq2 blocks in blue (Enrichment per bait/vector,
  Specificity, and compact Bait-vs-Vector, each pvalue+log2FoldChange),
  the Bayesian/MCMC block in purple (AdjEnr1/AdjEnr2 plus the four
  pairwise/specificity-adjusted probabilities), then a full per-dataset
  junction breakdown in gray (inframe_inorf/upstream/in_orf/downstream/
  in_frame/backwards/intron - the original 7 categories, not renamed here
  even though the Bait/Vector-stats blocks above use the newer
  "in-frame:Forward"/"out-of-frame:Forward" labels).

CRUSH

  A "crushed" prey pool is one where selection has driven most
  originally-present genes down to background/undetectable levels - the
  bait (or vector) successfully out-competed non-specific prey. The CRUSH
  table (above the Gene table) reports, separately for Vector1, Vector2,
  Bait1, and Bait2, what happened to each gene that started at >=3ppm in
  that dataset's own Non-Select sample (3ppm is the same detection floor
  used as the Bayesian model's threshold):

    High/Med/Low  - de-enriched by >=10x / 2.5x-10x / 1x-2.5x respectively
                    (High also includes anything that dropped below 3ppm
                    in Select, even if the fold-change itself is <10x)
    low/High Enrich - enriched by <2x / >=2x
    (these five are mutually exclusive and sum to 100% of that dataset's
    >=3ppm-in-Non-Select genes)

    Extinguished  - the old "casualty rate": % of those same genes that
                    fell below 3ppm in Select. This is a breakout of part
                    of "High", not a sixth exclusive bin, so it is not
                    part of the 100% sum above (shown grayed out).

  Good crush looks like Extinguished > ~80% - the large majority of genes
  that started present are gone by the end of selection. A weakly-crushed
  ("low-crush") dataset sits well below that.

STATISTICS METHOD (DESeq2)

  Adapted from the method described in:

    Velasquez-Zapata V, Elmore JM, Banerjee S, Dorman KS, Wise RP (2021)
    Next-generation yeast-two-hybrid analysis with Y2H-SCORES identifies
    novel interactors of the MLA immune receptor. PLOS Computational
    Biology 17(4): e1008890. https://doi.org/10.1371/journal.pcbi.1008890

  The enrichment contrast (Selected vs Non-Selected per bait/vector) is
  the same streamlined single-file approach Stat Maker v2 used. The
  specificity contrast (Bait1-selected vs Bait2-selected, when Bait2 is
  supplied) is a direct pairwise DESeq2 comparison - a scaled-down version
  of Y2H-SCORES' own specificity score (Software/spec_score.R), which
  generalizes to any number of baits via pairwise combinations; this
  build only ever has two baits, so there's only ever one pair to compare,
  and none of that generalization (or its kde2d-based fold-change
  reweighting) is needed here.

  The count matrix handed to DESeq2 is raw integer counts reconstructed
  from PPM (K = round(PPM * TotalReads / 1e6), TotalReads read from each
  gene_count_summary CSV's own header) - not PPM directly. Earlier versions
  fed PPM straight in with normalized=TRUE, which bypassed DESeq2's own
  size-factor normalization entirely (every sample's size factor forced to
  1). Stat Maker re-adds this as a distinct, parallel procedure: DESeq2's
  p-value computed on normalized PPM data (p-value_norm), included
  alongside the count-based result in the output.

  Stat Maker also lets DESeq2 estimate its own size factors via the
  "poscounts" method (estimateSizeFactors(dds, type="poscounts")), which
  computes each gene's normalization reference from only that gene's
  nonzero values - so a gene isn't excluded from the reference just because
  it happens to be zero in one or two samples, unlike DESeq2's classic
  median-of-ratios estimator (which requires nonzero everywhere).

  This matters because Y2H selection drives most genes toward zero in the
  Selected condition. Baits whose winning interactor doesn't fully
  out-compete background prey ("low-crush" datasets) previously showed a
  hit count that scaled with how incompletely the background was depleted,
  not with real interaction signal - poscounts normalization removes most
  of that effect (validated across a 73-dataset scan: correlation between
  a bait's casualty rate and its hit count at p<0.01 dropped from -0.73 to
  0.18, and the hit-count range compressed roughly 10-fold). The trade-off
  is a modestly higher empirical false-positive rate on true-null
  comparisons (~1.7% vs ~1.1% at nominal p<0.01). Most of these false
  positives are eliminated when an enrichment threshold of >1x is imposed,
  which ensures that only preys that increased their ppm under Select
  conditions are chosen.

  Stat Maker therefore reports the DESeq2 p-value computed two ways: on
  raw counts with poscounts size-factor normalization (p-value_raw), and
  on normalized PPM data as described in Y2H-SCORES (p-value_norm). With
  an enrichment threshold in place, the two converge.

  In practical terms: a low-crush bait (Extinguished well under ~80%, see
  CRUSH above) lets many non-specific prey survive selection just because
  the bait didn't out-compete them as hard as a well-crushed vector
  control would have - not because those prey are real interactors. Under
  the Bayesian/MCMC method this shows up directly as an inflated hit
  count, since AdjEnr/pBaitN_Vec compare Bait-Selected to Vector-Selected:
  those survivors look enriched relative to vector even though most of
  them are still de-enriched relative to their own gene's Non-Select PPM
  (just less de-enriched than in a strongly-crushed vector). DESeq2's
  poscounts normalization is what accounts for this when it computes a
  p-value - by estimating each gene's own size factor from its nonzero
  values instead of assuming uniform depletion, it stops rewarding a bait
  purely for having crushed less thoroughly than the vector it's compared
  against.

  Reported overdispersion is the mean overdispersion across the
  NON-SELECTED replicates only (a simple (Var-Mean)/Mean^2 estimate,
  computed before DESeq2 runs) - not DESeq2's own per-group dispersion
  estimate, which this program does not report.

STATISTICS METHOD (Bayesian/MCMC, via JAGS)

  This is DEEPN's original statistics method (Patrick Breheny,
  pbreheny/deepn), run here from an unmodified, vendored copy of that
  package (r_scripts/deepn_bayesian_src/) rather than re-derived. It
  requires JAGS 4.x installed separately (see "Verify R + JAGS
  Installation") - CRAN's rjags package hard-rejects JAGS 5.x.

  A per-gene negative-binomial hierarchical model is fit by MCMC
  (4 chains, 1500 adapt/burnin/sample iterations each - runMCMC() in the
  vendored mcmc.R). AdjEnr1/AdjEnr2 are posterior median log2 effect
  sizes (Bait1/Bait2 vs Vector) and reproduce almost exactly run-to-run.
  The p-columns (pBait1_Vec, pBait2_Vec, pBait1_Bait2, pBait2_Bait1,
  pBait1, pBait2) are pnorm(median/MAD) - a posterior probability, not a
  frequentist p-value - and because pnorm saturates hard toward 0/1, these
  columns can flip between runs for genes whose effect is only moderately
  confident, even though the underlying AdjEnr estimate barely moves. This
  is a known, inherent property of the original method (not something
  introduced in this port) - treat the p-columns as directional/qualitative
  and AdjEnr as the more stable quantity when the two seem to disagree.

  With just Bait1 + Vector supplied (no Bait2), the single-bait model
  (sModel1.jag) is used instead, reporting just AdjEnr/p.

R SCRIPT (run_Y2H_enrichement_stats_v5.R)
----------------------------------------------------------------------
"""


class MyListWidgetItem(QtWidgets.QListWidgetItem):
    def __init__(self, *args):
        super(MyListWidgetItem, self).__init__(*args)
        self.data = ''

    def GetData(self):
        return self.data


class RInstallThread(QtCore.QThread):
    lineOutput = QtCore.pyqtSignal(str)
    finished_ok = QtCore.pyqtSignal(bool, str)

    def __init__(self, rscript_exe, packages, parent=None):
        QtCore.QThread.__init__(self, parent)
        self.rscript_exe = rscript_exe
        self.packages = packages

    def run(self):
        try:
            sc.install_r_packages(self.rscript_exe, self.packages, progress_callback=self.lineOutput.emit)
            self.finished_ok.emit(True, '')
        except sc.RScriptError as e:
            self.finished_ok.emit(False, str(e))


class JagsDownloadThread(QtCore.QThread):
    lineOutput = QtCore.pyqtSignal(str)
    finished_ok = QtCore.pyqtSignal(bool, str)

    def run(self):
        try:
            pkg_path = sc.download_jags_pkg(progress_callback=self.lineOutput.emit)
            self.finished_ok.emit(True, pkg_path)
        except sc.RScriptError as e:
            self.finished_ok.emit(False, str(e))


class DeepnInstallThread(QtCore.QThread):
    lineOutput = QtCore.pyqtSignal(str)
    finished_ok = QtCore.pyqtSignal(bool, str)

    def __init__(self, rscript_exe, parent=None):
        QtCore.QThread.__init__(self, parent)
        self.rscript_exe = rscript_exe

    def run(self):
        try:
            sc.install_deepn_package(self.rscript_exe, progress_callback=self.lineOutput.emit)
            self.finished_ok.emit(True, '')
        except sc.RScriptError as e:
            self.finished_ok.emit(False, str(e))


class CollateThread(QtCore.QThread):
    lineOutput = QtCore.pyqtSignal(str)
    finished_ok = QtCore.pyqtSignal(bool, str)

    def __init__(self, parent_gui):
        QtCore.QThread.__init__(self, parent_gui)
        self.g = parent_gui

    def run(self):
        try:
            self.g._do_collate_and_stats(self.lineOutput.emit)
            self.finished_ok.emit(True, '')
        except Exception as e:
            self.finished_ok.emit(False, str(e))


# ---------- hit-criteria filtering/sorting (v6) ----------
# Pure functions (no self) - shared by the "Find" count preview and the
# Export_Sorted_B1/B2 workbook writer, so both always agree.

RAW_ENR_EPS = 0.05


def _raw_enrichment_ratio(p_sel, p_non):
    """Linear ppm select/non ratio. The numerator is used as-is - no
    pseudocount - so a 0-Select gene naturally comes out to ratio 0 (0 /
    anything positive is 0), rather than a symmetric pseudocount's
    misleading "unchanged" (ratio 1) for a gene that's actually absent.
    The denominator is floored at RAW_ENR_EPS rather than having it added:
    p_non below that floor is treated as RAW_ENR_EPS (avoids a division by
    a near-zero number blowing the ratio up), but any p_non already at or
    above it is used exactly as measured, not inflated by +0.05."""
    return p_sel / max(p_non, RAW_ENR_EPS)


# _gene_passes_criteria/_sort_key_value operate on a `maps` dict - {'p':
# {gene: val}, 'pval': {...}, 'infr': {...}, 'ratio': {...}, 'deseq2': {...},
# 'enr': {...}, 'adjenr': {...}} - one bait's worth of per-gene values,
# keyed the same way regardless of where they came from. This is what lets
# the hit-criteria panel work identically whether the data is a run that
# just finished (_build_bait_value_maps, computed from raw stats/bayesian_
# stats/ppm/bqp) or a previously-exported Stat Maker file re-opened from
# disk (_parse_sm_file, read straight back off that file's own columns -
# no raw source data needed).

def _gene_passes_criteria(gene, criteria, maps):
    """criteria: dict with keys 'p', 'pval_raw', 'pval_norm', 'infr',
    'adjenr_fold', 'deseq2_fold' - each None (or 0 for 'infr') means that
    threshold isn't set and never blocks a gene. Once a threshold IS set,
    a gene missing the underlying value fails it (can't confirm it clears
    a bar that isn't there).

    adjenr_fold/deseq2_fold are a dual enrichment-agreement check - AdjEnr
    (Bayesian/MCMC posterior median log2FC, Bait vs Vector) and DESeq2's
    own log2FC (Bait vs Vector) each independently need to clear their own
    threshold, rather than filtering on the plain within-bait select/non
    ratio. That raw ratio structurally misses "low-crush" hits (de-
    enriched within the bait's own culture, but still more enriched than
    Vector) - determined empirically by sweeping thresholds against the
    prior manually-curated Rab GTPase hit list (see rab_hit_recapitulation
    analysis, 2026-08-24): AdjEnr>0 alone recaptured 96.3% of that list at
    infr>=80% with far less noise than the raw ratio (76.2%) or DESeq2
    alone ever achieved in the same fold range; requiring DESeq2's log2FC
    to also clear >3x on top of AdjEnr>0 cut the noise further at
    negligible capture cost. Both are already log2-scale, so the
    threshold is compared directly, not converted from a linear ratio."""
    p = criteria.get('p')
    if p is not None:
        val = maps['p'].get(gene)
        if val is None or not (val > p):
            return False

    pval_raw = criteria.get('pval_raw')
    if pval_raw is not None:
        val = maps['pval_raw'].get(gene)
        if val is None or not (val <= pval_raw):
            return False

    pval_norm = criteria.get('pval_norm')
    if pval_norm is not None:
        val = maps['pval_norm'].get(gene)
        if val is None or not (val <= pval_norm):
            return False

    infr = criteria.get('infr') or 0
    if infr:
        val = maps['infr'].get(gene, 0)
        if not (val >= infr):
            return False

    adjenr_fold = criteria.get('adjenr_fold')
    if adjenr_fold is not None:
        val = maps['adjenr'].get(gene)
        if val is None or not (val > adjenr_fold):
            return False

    deseq2_fold = criteria.get('deseq2_fold')
    if deseq2_fold is not None:
        val = maps['deseq2'].get(gene)
        if val is None or not (val > deseq2_fold):
            return False

    return True


def _sort_key_value(gene, sort_key, maps):
    """Higher = more enriched, sorted descending. Missing values sink to
    the bottom of the passing block rather than erroring."""
    val = maps[sort_key].get(gene)
    return val if val is not None else float('-inf')


def _as_float_or_none(val):
    if val is None or val == 'NA' or val == '':
        return None
    return float(val)


def _default_junction_fn(bqp_data):
    """junction_fn(dataset_key, gene) -> the same dict get_junction_stats()
    returns, backed by real .bqp data - the fresh-run source."""
    return lambda dkey, gene: sc.get_junction_stats(bqp_data.get(dkey, {}), gene)


def _bait_vs_vector_keys(has_bait2):
    keys = ['pvalue_bait1_vs_vector', 'log2FoldChange_bait1_vs_vector']
    if has_bait2:
        keys += ['pvalue_bait2_vs_vector', 'log2FoldChange_bait2_vs_vector']
    return keys


def _suffix_bait_vs_vector(stats, has_bait2, suffix):
    """v7: DESeq2 now runs twice per collation - once on raw counts +
    poscounts (p-value_raw, the v5/v6 method) and once on normalized PPM
    data per Y2H-SCORES (p-value_norm, the pre-v5 method, re-added
    alongside rather than replaced). Only the Bait-vs-Vector contrast gets
    this raw/norm split (that's what the hit-criteria panel filters on);
    the within-culture Enrichment/Specificity contrasts stay single-
    sourced from the raw/poscounts run, unchanged. Renames
    pvalue_baitN_vs_vector/log2FoldChange_baitN_vs_vector to end in
    _raw/_norm in place."""
    keys = _bait_vs_vector_keys(has_bait2)
    for row in stats.values():
        for key in keys:
            if key in row:
                row[key + suffix] = row.pop(key)
    return stats


def _merge_bait_vs_vector_norm(stats, stats_norm, has_bait2):
    """Copies the _norm run's Bait-vs-Vector columns into `stats` (already
    _raw-suffixed by _suffix_bait_vs_vector) in place."""
    keys = _bait_vs_vector_keys(has_bait2)
    for gene, row_norm in stats_norm.items():
        row = stats.setdefault(gene, {})
        for key in keys:
            if key in row_norm:
                row[key + '_norm'] = row_norm[key]


def _build_bait_value_maps(bait_num, genes, ppm_data, stats, bayesian_stats, junction_fn):
    """Builds one bait's per-gene value maps for the hit-criteria panel.
    junction_fn(dataset_key, gene) -> percentage dict - either backed by
    real .bqp data (fresh run, _default_junction_fn) or by a loaded general
    .csv's own flat junction columns (_flat_junction_fn) - both produce
    identical maps either way."""
    suffix = 'bait%d' % bait_num
    sel_key, non_key = 'Bait%dS' % bait_num, 'Bait%dN' % bait_num
    maps = {k: {} for k in ('p', 'pval_raw', 'pval_norm', 'infr', 'ratio', 'deseq2', 'enr', 'adjenr')}
    for g in genes:
        brow = bayesian_stats.get(g, {})
        srow = stats.get(g, {})
        maps['p'][g] = _as_float_or_none(brow.get('pBait%d_Vec' % bait_num))
        maps['pval_raw'][g] = _as_float_or_none(srow.get('pvalue_%s_vs_vector_raw' % suffix))
        maps['pval_norm'][g] = _as_float_or_none(srow.get('pvalue_%s_vs_vector_norm' % suffix))
        maps['infr'][g] = junction_fn(sel_key, g)['in_frame']
        p_sel = ppm_data.get(sel_key, {}).get(g, 0.0)
        p_non = ppm_data.get(non_key, {}).get(g, 0.0)
        maps['ratio'][g] = _raw_enrichment_ratio(p_sel, p_non)
        maps['deseq2'][g] = _as_float_or_none(srow.get('log2FoldChange_%s_vs_vector_raw' % suffix))
        maps['enr'][g] = _as_float_or_none(brow.get('Enr%d' % bait_num))
        maps['adjenr'][g] = _as_float_or_none(brow.get('AdjEnr%d' % bait_num))
    return maps


# The general .csv - a flat dump of the exact same rich report
# _write_workbook produces as .xlsx (metadata, OVERDISPERSION, CRUSH
# table, three-row header, every column block), built by writing that
# .xlsx to a temp file and converting it row-by-row (see
# export_data_csv_clicked / sm6_batch.py's xlsx_to_csv). Reloading it
# (_parse_sm_csv) locates the same landmarks by their literal text
# ('Gene', 'OVERDISPERSION', 'Bait1 stats'/'Bait2 stats'/'Vector stats')
# rather than fixed row/column numbers, so it isn't fragile to a run
# having no Bait2 or the overdispersion rows differing in count.
#
# What survives the round-trip: everything in the Bait1/Bait2/Vector-
# stats blocks (ppm, raw enrichment, Enr/AdjEnr, DESeq2 log2fold/p-value,
# %in-frame:Forward) - full fidelity, this is the primary content and
# exactly what hit-criteria filtering needs. What does NOT survive: the
# old-report section's extra per-dataset junction categories (in_orf/
# upstream/downstream/backwards/intron - only in_frame is captured) and
# Vector1/Vector2's individual (pre-averaging) ppm. A workbook rebuilt
# from a loaded .csv therefore always passes include_old_report=False to
# _write_workbook (see _source_write_args) - omitting that section
# entirely rather than fabricating zeros for the categories that didn't
# round-trip. The original gene_count_summary/.bqp file paths aren't
# carried either - those can change machine to machine, and re-running
# the actual DESeq2/MCMC computation from a reloaded file isn't
# supported (see load_sm_file_clicked).

def _find_row(rows, col0_value):
    for i, row in enumerate(rows):
        if row and row[0] == col0_value:
            return i
    return None


def _find_col(row, value):
    for i, cell in enumerate(row):
        if cell == value:
            return i
    return None


def _to_float(v):
    if v in (None, ''):
        return None
    try:
        return float(v)
    except ValueError:
        return None


def _parse_sm_csv(path):
    with open(path, newline='') as fh:
        rows = list(csv.reader(fh))

    group_row_i = _find_row(rows, 'Gene')
    if group_row_i is None:
        raise ValueError("Couldn't find the 'Gene' header - is this a Stat Maker general .csv file?")
    field_row_i = group_row_i + 2
    field_row = rows[field_row_i]
    data_start = field_row_i + 1

    group_row = rows[group_row_i]
    bait1_col = _find_col(group_row, 'Bait1 stats')
    bait2_col = _find_col(group_row, 'Bait2 stats')
    vector_col = _find_col(group_row, 'Vector stats')
    if bait1_col is None or vector_col is None:
        raise ValueError("Couldn't find the Bait1/Vector stats blocks - is this a Stat Maker general .csv file?")
    has_bait2 = bait2_col is not None

    # Fixed field order within each block - see _write_workbook's `fields`/
    # `vector_fields` lists.
    BAIT_OFFSETS = {'ppm_non': 0, 'ppm_sel': 1, 'ratio': 2, 'enr': 3, 'adjenr': 4,
                    'deseq2': 5, 'p': 6, 'pval_raw': 7, 'pval_norm': 8,
                    'infr_sel': 9, 'infr_non': 10}
    VECTOR_OFFSETS = {'ppm_non': 0, 'ppm_sel': 1, 'ratio': 2, 'deseq2': 3, 'pval': 4,
                      'infr_sel': 5, 'infr_non': 6, 'outframe_non': 7, 'backward_non': 8}

    overdisp_i = _find_row(rows, 'OVERDISPERSION')
    overdispersion, bayes_overdispersion = None, {}
    if overdisp_i is not None:
        for r in rows[overdisp_i + 1:group_row_i]:
            if not r or len(r) < 2:
                continue
            label, val = r[0], _to_float(r[1])
            if val is None:
                continue
            if label == 'Vector non-select':
                bayes_overdispersion['vector_nonselect'] = val
            elif label == 'vector+Bait non-select':
                bayes_overdispersion['vector_bait_nonselect'] = val
            elif label == 'Vector select':
                bayes_overdispersion['vector_select'] = val
            elif label == 'DESeq Vector non-select':
                overdispersion = val

    genes, ppm_data = [], {k: {} for k in ('Bait1S', 'Bait1N', 'Bait2S', 'Bait2N',
                                           'Vector1S', 'Vector1N', 'Vector2S', 'Vector2N')}
    junction_flat, stats, bayesian_stats = {}, {}, {}
    for row in rows[data_start:]:
        if not row or not row[0]:
            continue
        g = row[0]
        genes.append(g)
        srow, brow = {}, {}

        def bait_vals(start_col, bait_num, suffix):
            get = lambda key: row[start_col + BAIT_OFFSETS[key]] if start_col + BAIT_OFFSETS[key] < len(row) else ''
            ppm_data['Bait%dN' % bait_num][g] = _to_float(get('ppm_non')) or 0.0
            ppm_data['Bait%dS' % bait_num][g] = _to_float(get('ppm_sel')) or 0.0
            brow['Enr%d' % bait_num] = get('enr')
            brow['AdjEnr%d' % bait_num] = get('adjenr')
            srow['log2FoldChange_%s_vs_vector_raw' % suffix] = get('deseq2')
            brow['pBait%d_Vec' % bait_num] = get('p')
            srow['pvalue_%s_vs_vector_raw' % suffix] = get('pval_raw')
            srow['pvalue_%s_vs_vector_norm' % suffix] = get('pval_norm')
            junction_flat[('Bait%dS' % bait_num, g)] = {'in_frame': _to_float(get('infr_sel')) or 0.0}
            junction_flat[('Bait%dN' % bait_num, g)] = {'in_frame': _to_float(get('infr_non')) or 0.0}

        bait_vals(bait1_col, 1, 'bait1')
        if has_bait2:
            bait_vals(bait2_col, 2, 'bait2')

        get_v = lambda key: row[vector_col + VECTOR_OFFSETS[key]] if vector_col + VECTOR_OFFSETS[key] < len(row) else ''
        v_non, v_sel = _to_float(get_v('ppm_non')) or 0.0, _to_float(get_v('ppm_sel')) or 0.0
        # Only the Vector1/Vector2 *average* survives the round-trip - used
        # for both replicate slots as the closest available approximation.
        for vk in ('Vector1N', 'Vector2N'):
            ppm_data[vk][g] = v_non
        for vk in ('Vector1S', 'Vector2S'):
            ppm_data[vk][g] = v_sel
        srow['log2FoldChange_vector'] = get_v('deseq2')
        srow['pvalue_vector'] = get_v('pval')
        infr_sel_v, infr_non_v = _to_float(get_v('infr_sel')) or 0.0, _to_float(get_v('infr_non')) or 0.0
        outframe_non_v = _to_float(get_v('outframe_non')) or 0.0
        backward_non_v = _to_float(get_v('backward_non')) or 0.0
        for vk in ('Vector1S', 'Vector2S'):
            junction_flat[(vk, g)] = {'in_frame': infr_sel_v}
        # Vector1/Vector2 aren't separately recoverable (the CSV only has
        # their average) - the same value in both slots, so averaging them
        # back in _write_workbook's Vector block reproduces it exactly.
        for vk in ('Vector1N', 'Vector2N'):
            junction_flat[(vk, g)] = {'in_frame': infr_non_v, 'not_in_frame': outframe_non_v,
                                      'backwards': backward_non_v}

        stats[g] = srow
        bayesian_stats[g] = brow

    def junction_fn(dkey, gene):
        return junction_flat.get((dkey, gene), {'in_frame': 0.0, 'not_in_frame': 0.0, 'backwards': 0.0})

    dataset_labels = {
        'Bait1N': 'Bait1_Non-Selected', 'Bait1S': 'Bait1_Selected',
        'Bait2N': 'Bait2_Non-Selected', 'Bait2S': 'Bait2_Selected',
        'Vector1N': 'Vector_Non-Selected_1', 'Vector2N': 'Vector_Non-Selected_2',
        'Vector1S': 'Vector_Selected_1', 'Vector2S': 'Vector_Selected_2',
    }
    dataset_keys = ['Bait1S', 'Bait1N'] + (['Bait2S', 'Bait2N'] if has_bait2 else []) + \
                  ['Vector1S', 'Vector1N', 'Vector2S', 'Vector2N']
    # The real paths are sitting in the file's own header rows (written by
    # _write_workbook as e.g. "Bait1_Selected,/Volumes/.../summary.csv") -
    # read them back by label rather than substituting a placeholder, so a
    # reloaded-then-re-exported workbook still names its actual inputs.
    label_to_key = {v: k for k, v in dataset_labels.items()}
    csv_paths = {}
    for row in rows[:group_row_i]:
        if len(row) >= 2 and row[0] in label_to_key:
            csv_paths[label_to_key[row[0]]] = row[1]
    for k in dataset_keys:
        csv_paths.setdefault(k, '(path not found in %s)' % os.path.basename(path))

    return {
        'name': os.path.splitext(os.path.basename(path))[0], 'genes': genes, 'has_bait2': has_bait2,
        'src_path': path, 'ppm_data': ppm_data, 'stats': stats, 'bayesian_stats': bayesian_stats,
        'junction_fn': junction_fn, 'csv_paths': csv_paths, 'dataset_labels': dataset_labels,
        'overdispersion': overdispersion, 'bayes_overdispersion': bayes_overdispersion,
    }


class ClickThroughSelector(QtWidgets.QWidget):
    """A label flanked by prev/next buttons, cycling through a fixed list
    of values (clamped at both ends, not wrapping). values[0] is always
    the "no filter" state; display_fn formats a value for the label."""

    changed = QtCore.pyqtSignal()

    def __init__(self, values, display_fn, parent=None):
        super(ClickThroughSelector, self).__init__(parent)
        self.values = values
        self.display_fn = display_fn
        self.index = 0

        layout = QtWidgets.QHBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        self.prev_btn = QtWidgets.QPushButton('<')
        self.next_btn = QtWidgets.QPushButton('>')
        self.prev_btn.setMinimumWidth(32)
        self.next_btn.setMinimumWidth(32)
        self.value_label = QtWidgets.QLabel()
        self.value_label.setAlignment(QtCore.Qt.AlignCenter)
        self.value_label.setMinimumWidth(80)
        layout.addWidget(self.prev_btn)
        layout.addWidget(self.value_label)
        layout.addWidget(self.next_btn)

        self.prev_btn.clicked.connect(self._prev)
        self.next_btn.clicked.connect(self._next)
        self._refresh()

    def _prev(self):
        if self.index > 0:
            self.index -= 1
            self._refresh()

    def _next(self):
        if self.index < len(self.values) - 1:
            self.index += 1
            self._refresh()

    def _refresh(self):
        self.value_label.setText(self.display_fn(self.values[self.index]))
        self.prev_btn.setEnabled(self.index > 0)
        self.next_btn.setEnabled(self.index < len(self.values) - 1)
        self.changed.emit()

    @property
    def value(self):
        return self.values[self.index]


class Stat_Maker_Gui(QtWidgets.QMainWindow, form_class):
    def __init__(self, *args):
        super(Stat_Maker_Gui, self).__init__(*args)
        self.setupUi(self)

        self.vec_sel_list.dropped.connect(self.vec_sel_fileDropped)
        self.vec_nonsel_list.dropped.connect(self.vec_nonsel_fileDropped)
        self.vec_sel_list_2.dropped.connect(self.vec_sel_fileDropped_2)
        self.vec_nonsel_list_2.dropped.connect(self.vec_nonsel_fileDropped_2)
        self.bait1_sel_list.dropped.connect(self.bait1_sel_fileDropped)
        self.bait1_nonsel_list.dropped.connect(self.bait1_nonsel_fileDropped)
        self.bait2_sel_list.dropped.connect(self.bait2_sel_fileDropped)
        self.bait2_nonsel_list.dropped.connect(self.bait2_nonsel_fileDropped)

        self.vec_sel_list.deleted.connect(self.file_deleted)
        self.vec_nonsel_list.deleted.connect(self.file_deleted)
        self.vec_sel_list_2.deleted.connect(self.file_deleted)
        self.vec_nonsel_list_2.deleted.connect(self.file_deleted)
        self.bait1_sel_list.deleted.connect(self.file_deleted)
        self.bait1_nonsel_list.deleted.connect(self.file_deleted)
        self.bait2_sel_list.deleted.connect(self.file_deleted)
        self.bait2_nonsel_list.deleted.connect(self.file_deleted)

        # Handler names deliberately do NOT match the on_<widget>_<signal>
        # pattern - setupUi()'s connectSlotsByName() binds a matching name to
        # BOTH of clicked's signal overloads (clicked() and clicked(bool)),
        # firing the handler twice per click even with no explicit connect
        # at all. Avoiding the naming convention and connecting explicitly,
        # once, is the only reliable way to get exactly one call per click.
        self.folder_choice_btn.clicked.connect(self.folder_choice_clicked)
        self.verify_env_btn.clicked.connect(self.verify_env_clicked)
        self.collate_btn.clicked.connect(self.collate_clicked)
        self.method_info_btn.clicked.connect(self.method_info_clicked)
        self.quit_btn.clicked.connect(self.quit_clicked)

        self.fileio = f.fileio()
        self.data = {}
        self.directory = '~/'
        self.rscript_exe = sc.find_rscript()
        self.jags_exe = sc.find_jags()
        self.jags_compatible = False
        if self.jags_exe:
            self.jags_compatible, self.jags_version = sc.check_jags_version(self.jags_exe)
        else:
            self.jags_version = None
        self.env_verified = False
        self.r_thread = None
        self.jags_download_thread = None
        self.deepn_thread = None
        self.collate_thread = None
        self._last_run = None
        # Whichever data source Export.csv/Find/Export_Sorted_B1/B2 all
        # operate on - {'kind': 'run', ...} for a just-completed run (real
        # bqp_data), or {'kind': 'file', ...} for a loaded general .csv
        # (flattened junction_fn instead) - see _source_write_args, which
        # is what makes _write_workbook work identically either way.
        self._active_source = None

        self._build_export_panel()
        self._build_results_panel()
        self.resize(self.width() + 560, int(self.height() * 1.20))

        self.log("Stat Maker v7 - collation + DESeq2 statistics (Bait1 required, Bait2 optional)")
        self.log("Found Rscript: %s" % self.rscript_exe if self.rscript_exe else
                 "Rscript not found on this machine yet.")
        if self.jags_exe and self.jags_compatible:
            self.log("Found jags %s (compatible): %s" % (self.jags_version, self.jags_exe))
        elif self.jags_exe:
            self.log("Found jags %s at %s, but rjags requires JAGS 4.x - this version won't work." %
                     (self.jags_version, self.jags_exe))
        else:
            self.log("jags not found on this machine yet.")
        if not (self.rscript_exe and self.jags_exe and self.jags_compatible):
            self.log("Click 'Verify R + JAGS Installation'.")

    def log(self, text):
        self.log_text.append(text)
        self.log_text.ensureCursorVisible()

    # ---------- R + JAGS environment check/install ----------
    # Three tiers, checked in order: R itself and JAGS itself are separate
    # standalone programs neither can install for the other - if missing,
    # this only ever points the user at a manual install (CRAN's own
    # installer for R; the official SourceForge installer for JAGS - NOT
    # Homebrew, whose `jags` formula is 5.0.0, a version CRAN's rjags
    # package hard-rejects at its own configure step). Once both exist,
    # the R *packages* each needs (DESeq2 etc., and now rjags/runjags) are
    # auto-installable through R's own package manager, same as always.

    @QtCore.pyqtSlot()
    def verify_env_clicked(self):
        self.verify_env_btn.setEnabled(False)
        self._verify_rscript_step()

    def _verify_rscript_step(self):
        self.rscript_exe = sc.find_rscript()
        if self.rscript_exe is None:
            box = QtWidgets.QMessageBox(self)
            box.setWindowTitle("R Not Found")
            box.setText("R does not appear to be installed on this Mac.\n\n"
                        "DEEPN's statistics step needs R (with DESeq2) to run.")
            download_btn = box.addButton("Download R...", QtWidgets.QMessageBox.ActionRole)
            box.addButton("Cancel", QtWidgets.QMessageBox.RejectRole)
            box.exec_()
            if box.clickedButton() == download_btn:
                QtGui.QDesktopServices.openUrl(QtCore.QUrl(sc.CRAN_DOWNLOAD_URL))
                self.log("Opened %s - after installing R, click 'Verify R + JAGS Installation' again." % sc.CRAN_DOWNLOAD_URL)
            self.verify_env_btn.setEnabled(True)
            return

        self.log("Rscript found: %s" % self.rscript_exe)
        self._verify_jags_step()

    def _verify_jags_step(self):
        self.jags_exe = sc.find_jags()
        self.jags_compatible = False
        self.jags_version = None
        if self.jags_exe:
            self.jags_compatible, self.jags_version = sc.check_jags_version(self.jags_exe)

        if not self.jags_compatible:
            box = QtWidgets.QMessageBox(self)
            box.setWindowTitle("JAGS 4.x Required")
            if self.jags_exe:
                box.setText(
                    "Found JAGS %s at %s, but the Bayesian statistics method needs "
                    "JAGS 4.x specifically - rjags (the R/JAGS bridge) rejects JAGS 5.x.\n\n"
                    "Note: Homebrew's 'jags' formula currently only builds 5.x and is NOT "
                    "usable here. Install the official JAGS 4.3.2 package below instead."
                    % (self.jags_version, self.jags_exe))
            else:
                box.setText(
                    "JAGS (the MCMC engine the Bayesian statistics method needs) does not "
                    "appear to be installed on this Mac.\n\n"
                    "Install JAGS 4.3.2 from the official installer below - NOT Homebrew's "
                    "'jags' formula, which is version 5.x and incompatible with rjags.")
            fetch_btn = box.addButton("Download and Open Installer", QtWidgets.QMessageBox.ActionRole)
            download_btn = box.addButton("Open Download Page", QtWidgets.QMessageBox.ActionRole)
            box.addButton("Cancel", QtWidgets.QMessageBox.RejectRole)
            box.exec_()
            clicked = box.clickedButton()
            if clicked == fetch_btn:
                self.log("Downloading %s ..." % sc.JAGS_PKG_URL)
                self.jags_download_thread = JagsDownloadThread(self)
                self.jags_download_thread.lineOutput.connect(self.log)
                self.jags_download_thread.finished_ok.connect(self._jags_download_finished)
                self.jags_download_thread.start()
                return
            elif clicked == download_btn:
                QtGui.QDesktopServices.openUrl(QtCore.QUrl(sc.JAGS_DOWNLOAD_URL))
                self.log("Opened %s - after installing JAGS 4.x, click 'Verify R + JAGS Installation' again." % sc.JAGS_DOWNLOAD_URL)
            self.verify_env_btn.setEnabled(True)
            return

        self.log("jags %s found (compatible): %s" % (self.jags_version, self.jags_exe))
        self._verify_packages_step()

    @QtCore.pyqtSlot(bool, str)
    def _jags_download_finished(self, ok, result):
        if not ok:
            self.log("JAGS download FAILED: %s" % result)
            QtWidgets.QMessageBox.critical(self, "Download Failed", result[:2000])
            self.verify_env_btn.setEnabled(True)
            return
        pkg_path = result
        self.log("Opening %s in Installer.app - follow the prompts and enter your Mac "
                 "password when asked, then click 'Verify R + JAGS Installation' again." % pkg_path)
        sc.open_installer(pkg_path)
        self.verify_env_btn.setEnabled(True)

    def _verify_packages_step(self):
        self.log("Checking required R packages...")
        QtWidgets.QApplication.processEvents()
        packages = sc.REQUIRED_R_PACKAGES + sc.BAYESIAN_R_PACKAGES
        status = sc.check_r_packages(self.rscript_exe, packages)
        missing = [p for p, ok in status.items() if not ok]
        for p, ok in status.items():
            self.log("  %-10s %s" % (p, "OK" if ok else "MISSING"))

        if not missing:
            self._verify_deepn_step()
            return

        reply = QtWidgets.QMessageBox.question(
            self, "Install R Packages?",
            "The following R packages need to be installed:\n\n%s\n\n"
            "This can take several minutes (downloading/compiling DESeq2 and its dependencies). Proceed?"
            % ", ".join(missing),
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No
        )
        if reply != QtWidgets.QMessageBox.Yes:
            self.verify_env_btn.setEnabled(True)
            return

        self.log("Installing: %s ..." % ", ".join(missing))
        self.r_thread = RInstallThread(self.rscript_exe, missing, self)
        self.r_thread.lineOutput.connect(self.log)
        self.r_thread.finished_ok.connect(self._r_install_finished)
        self.r_thread.start()

    @QtCore.pyqtSlot(bool, str)
    def _r_install_finished(self, ok, err):
        if not ok:
            self.log("R package installation FAILED: %s" % err)
            QtWidgets.QMessageBox.critical(self, "Installation Failed", err[:2000])
            self.verify_env_btn.setEnabled(True)
            return
        self.log("R package installation finished. Re-checking...")
        packages = sc.REQUIRED_R_PACKAGES + sc.BAYESIAN_R_PACKAGES
        status = sc.check_r_packages(self.rscript_exe, packages)
        missing = [p for p, ok2 in status.items() if not ok2]
        if missing:
            self.log("Still missing: %s" % ", ".join(missing))
            self.verify_env_btn.setEnabled(True)
        else:
            self._verify_deepn_step()

    def _verify_deepn_step(self):
        self.log("Checking vendored 'deepn' package (pbreheny/deepn, hardwired local copy)...")
        QtWidgets.QApplication.processEvents()
        if sc.check_deepn_package(self.rscript_exe):
            self.log("  deepn      OK")
            self._env_fully_verified()
            return

        self.log("  deepn      MISSING - installing from vendored local source...")
        self.deepn_thread = DeepnInstallThread(self.rscript_exe, self)
        self.deepn_thread.lineOutput.connect(self.log)
        self.deepn_thread.finished_ok.connect(self._deepn_install_finished)
        self.deepn_thread.start()

    @QtCore.pyqtSlot(bool, str)
    def _deepn_install_finished(self, ok, err):
        if not ok:
            self.log("deepn package installation FAILED: %s" % err)
            QtWidgets.QMessageBox.critical(self, "Installation Failed", err[:2000])
            self.verify_env_btn.setEnabled(True)
            return
        self.log("deepn package installed.")
        self._env_fully_verified()

    def _env_fully_verified(self):
        self.env_verified = True
        self.log("R + JAGS environment fully verified.")
        self.verify_env_btn.setText("R + JAGS Verified")
        self.verify_env_btn.setEnabled(False)
        self._update_collate_enabled()

    # ---------- folder / file selection ----------

    def initialize_folders(self, directory):
        self.fileio.create_new_folder(directory, "stat_maker_output")

    @QtCore.pyqtSlot()
    def folder_choice_clicked(self):
        directory = str(QtWidgets.QFileDialog.getExistingDirectory(QtWidgets.QFileDialog(), "Locate Work Folder",
                                                               os.path.expanduser("~"),
                                                               QtWidgets.QFileDialog.ShowDirsOnly))
        if not directory:
            return
        self.directory = directory
        self.initialize_folders(self.directory)
        self.load_gene_summary_files()

    def load_gene_summary_files(self):
        try:
            dirlist = os.listdir(os.path.join(self.directory, 'gene_count_summary'))
            self.file_list.clear()
            for file in sorted(dirlist):
                if not re.match(r'^\.', file) and re.match(r'.+summary\.csv', file):
                    path = os.path.join(self.directory, 'gene_count_summary', file)
                    iconProvider = QtWidgets.QFileIconProvider()
                    fileInfo = QtCore.QFileInfo(path)
                    icon = iconProvider.icon(fileInfo)
                    item = MyListWidgetItem()
                    item.data = path
                    item.setIcon(icon)
                    item.setText(file)
                    self.file_list.addItem(item)
        except OSError:
            pass

    def fileDropped(self, list_widget, url, key):
        if os.path.exists(url):
            iconProvider = QtWidgets.QFileIconProvider()
            fileInfo = QtCore.QFileInfo(url)
            icon = iconProvider.icon(fileInfo)
            item = MyListWidgetItem()
            item.data = url
            item.setIcon(icon)
            item.setText(os.path.basename(os.path.normpath(url)))
            list_widget.clear()
            list_widget.addItem(item)
            self.data[key] = url
        else:
            self.data.pop(key, None)
        self._update_collate_enabled()

    def file_deleted(self, path):
        for key in list(self.data.keys()):
            if self.data[key] == path:
                del self.data[key]
        self._update_collate_enabled()

    def vec_sel_fileDropped(self, path):
        self.fileDropped(self.vec_sel_list, str(path), 'Vector1S')

    def vec_nonsel_fileDropped(self, path):
        self.fileDropped(self.vec_nonsel_list, str(path), 'Vector1N')

    def vec_sel_fileDropped_2(self, path):
        self.fileDropped(self.vec_sel_list_2, str(path), 'Vector2S')

    def vec_nonsel_fileDropped_2(self, path):
        self.fileDropped(self.vec_nonsel_list_2, str(path), 'Vector2N')

    def bait1_sel_fileDropped(self, path):
        self.fileDropped(self.bait1_sel_list, str(path), 'Bait1S')

    def bait1_nonsel_fileDropped(self, path):
        self.fileDropped(self.bait1_nonsel_list, str(path), 'Bait1N')

    def bait2_sel_fileDropped(self, path):
        self.fileDropped(self.bait2_sel_list, str(path), 'Bait2S')

    def bait2_nonsel_fileDropped(self, path):
        self.fileDropped(self.bait2_nonsel_list, str(path), 'Bait2N')

    def _has_bait2(self):
        return 'Bait2S' in self.data and 'Bait2N' in self.data

    def _update_collate_enabled(self):
        required = ('Vector1S', 'Vector1N', 'Vector2S', 'Vector2N', 'Bait1S', 'Bait1N')
        ready = all(k in self.data for k in required)
        # Bait2 is optional, but if only one of its two files is dropped
        # (not both), that's a half-finished state - don't run until it's
        # either both-present or both-absent.
        bait2_keys = ('Bait2S', 'Bait2N')
        bait2_partial = any(k in self.data for k in bait2_keys) and not all(k in self.data for k in bait2_keys)
        self.collate_btn.setEnabled(ready and not bait2_partial)

    # ---------- collate + stats ----------

    @QtCore.pyqtSlot()
    def collate_clicked(self):
        if not self.env_verified:
            reply = QtWidgets.QMessageBox.question(
                self, "Environment Not Verified",
                "R + JAGS installation hasn't been verified yet this session. Continue anyway "
                "(the statistics step will fail if R/DESeq2/JAGS isn't actually available)?",
                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
            if reply != QtWidgets.QMessageBox.Yes:
                return

        self.collate_btn.setEnabled(False)
        self.statusbar.showMessage("Running collation + statistics...")
        self.collate_thread = CollateThread(self)
        self.collate_thread.lineOutput.connect(self.log)
        self.collate_thread.finished_ok.connect(self._collate_finished)
        self.collate_thread.start()

    @QtCore.pyqtSlot(bool, str)
    def _collate_finished(self, ok, err):
        self._update_collate_enabled()
        if ok:
            self.statusbar.showMessage("Done.", 8000)
            run = self._last_run
            junction_fn = _default_junction_fn(run['bqp_data'])
            self._active_source = {
                'kind': 'run', 'genes': run['genes'], 'has_bait2': run['has_bait2'],
                'out_folder': run['out_folder'], 'ppm_data': run['ppm_data'],
                'maps1': _build_bait_value_maps(1, run['genes'], run['ppm_data'], run['stats'],
                                                run['bayesian_stats'], junction_fn),
                'maps2': (_build_bait_value_maps(2, run['genes'], run['ppm_data'], run['stats'],
                                                 run['bayesian_stats'], junction_fn)
                         if run['has_bait2'] else None),
            }
            self.export_default_btn.setEnabled(True)
            self.export_data_btn.setEnabled(True)
            self.b1_criteria['find_btn'].setEnabled(True)
            self.b1_criteria['export_btn'].setEnabled(True)
            self.b2_criteria['find_btn'].setEnabled(run['has_bait2'])
            self.b2_criteria['export_btn'].setEnabled(run['has_bait2'])
            self.bait2_criteria_group.setVisible(run['has_bait2'])
            self.bait2_results_group.setVisible(run['has_bait2'])
            self.b1_count_label.setText('')
            self.b2_count_label.setText('')
            self.b1_results_table.setRowCount(0)
            self.b2_results_table.setRowCount(0)
            self.source_label.setText("Active source: just-completed run")
            self.export_name_edit.setText(datetime.now().strftime('stat_maker_%d%b%Y-%H%M%S'))
        else:
            self.statusbar.showMessage("Failed - see log.", 8000)
            self.log("ERROR: %s" % err)
            QtWidgets.QMessageBox.critical(self, "Collation Failed", err[:2000])

    def _do_collate_and_stats(self, log_cb):
        directory = self.directory
        out_folder = os.path.join(directory, 'stat_maker_output')
        self.fileio.create_new_folder(directory, 'stat_maker_output')

        has_bait2 = self._has_bait2()
        keys = ['Vector1S', 'Vector1N', 'Vector2S', 'Vector2N', 'Bait1S', 'Bait1N']
        if has_bait2:
            keys += ['Bait2S', 'Bait2N']
        csv_paths = {k: self.data[k] for k in keys}

        log_cb("Checking gene counts are consistent across all %d files..." % len(csv_paths))
        consistent, counts = sc.check_gene_counts_consistent(list(csv_paths.values()))
        for k, path in csv_paths.items():
            log_cb("  %-10s %6d genes  (%s)" % (k, counts[path], os.path.basename(path)))
        if not consistent:
            log_cb("WARNING: gene counts differ across files - they may not have been "
                   "processed with the same gene dictionary/list.")

        log_cb("Loading raw counts (PPM reconstructed to integer counts via TotalReads, "
               "summed across scaffold-duplicate gene rows)...")
        raw_count_data = {k: sc.load_raw_counts_summed(path) for k, path in csv_paths.items()}

        timestamp = datetime.now().strftime('%d%b%Y-%H%M%S')
        run_out_dir = os.path.join(out_folder, 'run_%s' % timestamp)
        os.makedirs(run_out_dir, exist_ok=True)
        collated_path = os.path.join(run_out_dir, 'collated_input.csv')
        sc.build_collated_input_v5(raw_count_data, run_out_dir, collated_path, has_bait2)
        log_cb("Wrote %s%s" % (collated_path, " (Bait1 + Bait2 + Vector)" if has_bait2 else " (Bait1 + Vector only)"))

        log_cb("Running DESeq2 statistics on raw counts + poscounts normalization (p-value_raw)...")
        r_output, overdispersion = sc.run_r_script(R_SCRIPT_PATH, collated_path, self.rscript_exe)
        for line in r_output.splitlines():
            log_cb("  [R] " + line)
        if overdispersion is not None:
            log_cb("Non-selected overdispersion: %.4f" % overdispersion)

        stats_csv = os.path.join(run_out_dir, 'everything_combined.csv')
        stats = sc.load_stats_output(stats_csv)
        stats = _suffix_bait_vs_vector(stats, has_bait2, '_raw')

        log_cb("Running DESeq2 statistics again on normalized PPM data, as described in "
               "Y2H-SCORES (p-value_norm)...")
        ppm_data_for_norm = {k: sc.load_ppm_summed(path) for k, path in csv_paths.items()}
        norm_run_dir = os.path.join(run_out_dir, 'norm')
        os.makedirs(norm_run_dir, exist_ok=True)
        collated_path_norm = os.path.join(norm_run_dir, 'collated_input.csv')
        sc.build_collated_input_v3(ppm_data_for_norm, norm_run_dir, collated_path_norm, has_bait2)
        r_output_norm, _ = sc.run_r_script(R_SCRIPT_PATH, collated_path_norm, self.rscript_exe)
        for line in r_output_norm.splitlines():
            log_cb("  [R norm] " + line)
        stats_norm = sc.load_stats_output(os.path.join(norm_run_dir, 'everything_combined.csv'))
        _merge_bait_vs_vector_norm(stats, stats_norm, has_bait2)

        log_cb("Running Bayesian/MCMC statistics (analyzeDeepn via JAGS - this is the slow "
               "step, several minutes)...")
        try:
            bayes_output, bayes_csv, bayes_overdispersion = sc.run_bayesian_r_script(
                BAYESIAN_R_SCRIPT_PATH, csv_paths, BAYESIAN_THRESHOLD_PPM, run_out_dir, self.rscript_exe)
            for line in bayes_output.splitlines():
                log_cb("  [R/JAGS] " + line)
            bayesian_stats = sc.load_bayesian_stats_output(bayes_csv)
        except sc.RScriptError as e:
            log_cb("Bayesian/MCMC statistics FAILED (DESeq2 results above are still valid): %s" % e)
            bayesian_stats = {}
            bayes_overdispersion = {}

        log_cb("Loading .bqp junction data for collation...")
        bqp_data = {}
        for k, path in csv_paths.items():
            basename = os.path.basename(path).replace('_summary.csv', '')
            bqp_path = sc.find_bqp_path(directory, basename)
            if bqp_path:
                if os.path.basename(bqp_path) == basename + '.bqp':
                    log_cb("  %-10s using legacy-named 5p file (no prefix): %s" % (k, os.path.basename(bqp_path)))
                bqp_data[k] = sc.load_bqp(bqp_path)
            else:
                log_cb("  WARNING: no .bqp found for %s (checked 5p_%s.bqp and %s.bqp) - junction stats will be 0 for it"
                       % (k, basename, basename))
                bqp_data[k] = {}

        dataset_labels = {
            'Bait1N': 'Bait1_Non-Selected', 'Bait1S': 'Bait1_Selected',
            'Bait2N': 'Bait2_Non-Selected', 'Bait2S': 'Bait2_Selected',
            'Vector1N': 'Vector_Non-Selected_1', 'Vector2N': 'Vector_Non-Selected_2',
            'Vector1S': 'Vector_Selected_1', 'Vector2S': 'Vector_Selected_2',
        }

        genes = sorted(raw_count_data['Bait1S'].keys())
        log_cb("Collating junction stats for %d genes..." % len(genes))

        # Cached so the Export/Find/Export_Sorted_* actions in the hit-
        # criteria panel can reuse this run's results without re-running
        # DESeq2/MCMC (which is the expensive, several-minutes part).
        self._last_run = {
            'out_folder': out_folder, 'csv_paths': csv_paths, 'dataset_labels': dataset_labels,
            'overdispersion': overdispersion, 'bayes_overdispersion': bayes_overdispersion,
            'stats': stats, 'bayesian_stats': bayesian_stats, 'bqp_data': bqp_data,
            'genes': genes, 'has_bait2': has_bait2,
            'ppm_data': {k: sc.load_ppm_summed(path) for k, path in csv_paths.items()},
        }

        out_name = 'stat_maker_%s.xlsx' % timestamp
        out_path = os.path.join(out_folder, out_name)
        self._write_workbook(out_path, csv_paths, dataset_labels, overdispersion,
                             bayes_overdispersion, stats, bayesian_stats, bqp_data, genes, has_bait2, log_cb,
                             ppm_data=self._last_run['ppm_data'])
        log_cb("Wrote %s" % out_path)

    # ---------- output workbook ----------
    # Two-row column headers: an upper "group" row (merged, color-coded by
    # method - DESeq2 now, matching colors reserved for an MCMC/Bayesian
    # block once that method is added) sitting above the per-field header
    # row that the Excel autofilter buttons (sort/filter dropdowns, like a
    # pivot table) actually attach to.

    def _write_workbook(self, out_path, csv_paths, dataset_labels, overdispersion,
                        bayes_overdispersion, stats, bayesian_stats, bqp_data, genes, has_bait2, log_cb,
                        sort_spec=None, criteria=None, ppm_data=None, junction_fn=None, include_old_report=True):
        """sort_spec/criteria: both None for the plain, unsorted/unfiltered
        export. Otherwise sort_spec=(bait_num, sort_key) with sort_key in
        ('deseq2','enr','adjenr'), and criteria=that bait's threshold dict
        (see _gene_passes_criteria) - genes passing all of criteria's set
        thresholds are pulled to the top, sorted descending by sort_key,
        and shaded light orange; every other gene follows in its existing
        (alphabetical) order, unshaded.

        junction_fn(dataset_key, gene) -> get_junction_stats()-shaped dict:
        defaults to reading real bqp_data, but a workbook rebuilt from a
        loaded general .csv (which only has flattened junction
        percentages, not raw .bqp objects) supplies its own instead - see
        _read_general_csv."""
        if junction_fn is None:
            junction_fn = _default_junction_fn(bqp_data)
        workbook = xls.Workbook(out_path)
        ws = workbook.add_worksheet('results')

        # Old-report block colors (DESeq2 blue / Bayesian purple / junction
        # gray) - unchanged from before.
        fmt_group_deseq2 = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter',
                                                'bg_color': '#D9E8FB', 'border': 1})
        fmt_group_bayesian = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter',
                                                   'bg_color': '#E6D9F7', 'border': 1})
        fmt_group_junction = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter',
                                                   'bg_color': '#E9E9E9', 'border': 1})
        # Bait1/Bait2/Vector-stats block colors, matched to Example_For
        # statmaker.xlsx's own theme-color header fills (extracted from
        # that file's xl/theme/theme1.xml: theme4/accent4 "Gold, Accent 4,
        # Lighter 60%", theme2/accent2 "Orange, Accent 2, Lighter 60%",
        # theme0/lt1 "White, Background 1, Darker 15%").
        fmt_group_bait1 = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter',
                                               'bg_color': '#FFE699', 'border': 1})
        fmt_group_bait2 = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter',
                                               'bg_color': '#F8CBAD', 'border': 1})
        fmt_group_vector = workbook.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter',
                                                'bg_color': '#D9D9D9', 'border': 1})
        fmt_field = workbook.add_format({'bold': True, 'border': 1})
        # Per-column source-method tint within the Bait1/Bait2/Vector-stats
        # blocks: Bayesian-derived columns (Enr/AdjEnr/pBaitN_Vec) purple,
        # DESeq2-derived columns (log2fold/p-value) blue - same hues as the
        # old-report block colors, just applied per-cell instead of merged
        # across a whole block. The Select junction-% column gets its own
        # pink tint; the Non-Select one stays uncolored.
        fmt_field_bayesian = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#E6D9F7'})
        fmt_field_deseq2 = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#D9E8FB'})
        fmt_field_select = workbook.add_format({'bold': True, 'border': 1, 'bg_color': '#EEB1C2'})
        fmt_unit = workbook.add_format({'italic': True, 'align': 'center', 'font_color': '#666666'})
        fmt_gene = workbook.add_format({'bold': True})
        fmt_section = workbook.add_format({'bold': True})
        # Extinguished is a breakout/subset stat, not one of the 5 mutually
        # exclusive CRUSH bins - grayed out (same gray as Vector stats) to
        # set it apart visually.
        fmt_gray = workbook.add_format({'bg_color': '#D9D9D9'})
        fmt_gray_bold = workbook.add_format({'bold': True, 'bg_color': '#D9D9D9'})
        # Display-only cap at 2 decimal places for every numeric gene-stat
        # cell (CRUSH percentages and the whole per-gene table) - the
        # underlying value written to the cell keeps full precision, only
        # the displayed text is rounded.
        fmt_num2 = workbook.add_format({'num_format': '0.00'})
        fmt_num2_gray = workbook.add_format({'num_format': '0.00', 'bg_color': '#D9D9D9'})
        # Full-row highlight for genes passing a hit-criteria filter
        # (sort_spec/criteria below) - substituted in place of fmt_num2/
        # fmt_gene for every cell in a passing gene's row.
        fmt_num2_orange = workbook.add_format({'num_format': '0.00', 'bg_color': '#FCE4C0'})
        fmt_gene_orange = workbook.add_format({'bold': True, 'bg_color': '#FCE4C0'})

        row = 0
        for k in csv_paths:
            ws.write(row, 0, dataset_labels[k])
            ws.write(row, 1, csv_paths[k])
            row += 1

        # Both methods' overdispersion estimates, side by side for direct
        # comparison: analyzeDeepn()'s edgeR-based estimates (the original
        # Bayesian-method math, computed on raw counts) alongside DESeq2's
        # own estimate (computed on PPM) - deliberately kept separate since
        # they're different quantities from different pipelines, not
        # expected to match each other.
        overdisp_rows = [
            ('Vector non-select', bayes_overdispersion.get('vector_nonselect')),
            ('vector+Bait non-select', bayes_overdispersion.get('vector_bait_nonselect')),
            ('Vector select', bayes_overdispersion.get('vector_select')),
            ('DESeq Vector non-select', overdispersion),
        ]
        if any(v is not None for _, v in overdisp_rows):
            ws.write(row, 0, 'OVERDISPERSION', fmt_section)
            row += 1
            for label, value in overdisp_rows:
                ws.write(row, 0, label)
                if value is not None:
                    ws.write_number(row, 1, value)
                row += 1

        # PPM for every dataset key, loaded once (not per-gene) - unless
        # the caller already has it cached (Export_Sorted_*/Export.csv all
        # reuse the same run's ppm_data instead of reloading every click).
        if ppm_data is None:
            ppm_data = {k: sc.load_ppm_summed(path) for k, path in csv_paths.items()}

        # ---- CRUSH table: how hard each individual Sel/Non pair "crushed"
        # its own prey pool by itself, independent of any bait-vs-vector
        # comparison. Restricted to genes that started >=3ppm in Non-Select
        # (the same detection floor used as the Bayesian threshold) - below
        # that, a gene was never reliably present to begin with.
        #   High:       ppm_select < 3 (dropped below detection) OR
        #               select/non-select ratio <= 0.1 (>=10-fold lost)
        #   Med:        not High, ratio in (0.1, 0.4]  (2.5x-10x lost)
        #   Low:        not High, ratio in (0.4, 1)    (1x-2.5x lost)
        #   low Enrich: ratio in [1, 2)
        #   High Enrich: ratio >= 2
        # (High/Med/Low/low-Enrich/High-Enrich are mutually exclusive and
        # sum to 100% of the >=3ppm-in-Non-Select denominator.)
        #   Extinguished: ppm_select < 3 alone (the old casualty rate) - a
        # breakout of part of "High", not a 6th exclusive bin, so it does
        # not add into that 100% sum.
        CRUSH_THRESHOLD = 3.0

        def crush_bins(sel_key, non_key):
            ppm_sel = ppm_data.get(sel_key, {})
            ppm_non = ppm_data.get(non_key, {})
            present = [g for g in ppm_non if ppm_non[g] >= CRUSH_THRESHOLD]
            if not present:
                return None
            counts = {'high': 0, 'med': 0, 'low_de': 0, 'low_en': 0, 'high_en': 0, 'extinguished': 0}
            for g in present:
                s = ppm_sel.get(g, 0.0)
                ratio = s / ppm_non[g]
                if s < CRUSH_THRESHOLD:
                    counts['extinguished'] += 1
                if s < CRUSH_THRESHOLD or ratio <= 0.1:
                    counts['high'] += 1
                elif ratio <= 0.4:
                    counts['med'] += 1
                elif ratio < 1:
                    counts['low_de'] += 1
                elif ratio < 2:
                    counts['low_en'] += 1
                else:
                    counts['high_en'] += 1
            n = len(present)
            return {k: 100.0 * v / n for k, v in counts.items()}

        crush_rows = [('Vector1', 'Vector1S', 'Vector1N'), ('Vector2', 'Vector2S', 'Vector2N'),
                      ('Bait1', 'Bait1S', 'Bait1N')]
        if has_bait2:
            crush_rows.append(('Bait2', 'Bait2S', 'Bait2N'))

        ws.write(row, 0, 'CRUSH', fmt_section)
        ws.write(row, 1, 'De-Enrichment fold (percent)')
        ws.write(row, 4, 'Enrichment fold (Percent)')
        ws.write(row, 6, 'Percent', fmt_gray_bold)
        row += 1
        for c, label in ((1, 'High (>=10)'), (2, 'Med (2.5-10)'), (3, 'Low (1-2.5)'), (4, 'low (<2)'), (5, 'High (>=2)')):
            ws.write(row, c, label)
        ws.write(row, 6, 'Extinguished', fmt_gray_bold)
        row += 1
        for label, sel_key, non_key in crush_rows:
            ws.write(row, 0, label)
            bins = crush_bins(sel_key, non_key)
            if bins is not None:
                for c, key in ((1, 'high'), (2, 'med'), (3, 'low_de'), (4, 'low_en'), (5, 'high_en')):
                    ws.write_number(row, c, bins[key], fmt_num2)
                ws.write_number(row, 6, bins['extinguished'], fmt_num2_gray)
            row += 1

        # Three header rows: group_row (top) titles only the Bait1/Bait2/
        # Vector-stats blocks; mid_row carries per-column scale/unit tags
        # for those same blocks (Linear FC / log2 / P / p-value / Percent)
        # and doubles as the block-title row for the appended old-report
        # section further right (which has nothing on group_row); field_row
        # has the actual column names for everything, and is what the
        # autofilter (sort/filter dropdowns) attaches to.
        group_row = row
        mid_row = row + 1
        field_row = row + 2

        # Gene column spans all three header rows
        ws.merge_range(group_row, 0, field_row, 0, 'Gene', fmt_gene)

        def vector_mean_ppm(gene, cond):
            """Mean of Vector1<cond>/Vector2<cond> PPM for one gene, cond='S' or 'N'."""
            vals = [ppm_data[k].get(gene, 0.0) for k in ('Vector1' + cond, 'Vector2' + cond) if k in ppm_data]
            return sum(vals) / len(vals) if vals else 0.0

        # See _raw_enrichment_ratio - denominator-only pseudocount, unlike
        # Enr1's symmetric one, so 0-Select naturally comes out to ratio 0.

        col = 1
        bait_blocks = [('Bait1', 'bait1', 'Enr1', 'AdjEnr1', 'pBait1_Vec')]
        if has_bait2:
            bait_blocks.append(('Bait2', 'bait2', 'Enr2', 'AdjEnr2', 'pBait2_Vec'))

        # kind: 'ppm_non' / 'ppm_sel' / 'raw_enr' / 'bayesian' / 'deseq2' / 'junction'
        # extra: bayesian dict key / stats dict key / (bqp_key, junction_stat_key)
        # unit: scale/type tag written on mid_row (None -> left blank)
        # fld_fmt: per-field-cell format override (None -> fmt_field)
        bait_field_cols = []  # (bait_label, kind, extra, worksheet_col)
        for bait_label, suffix, enr_key, adjenr_key, pvec_key in bait_blocks:
            start_col = col
            fields = [
                ('ppm non', 'ppm_non', None, None, None),
                ('ppm select', 'ppm_sel', None, None, None),
                ('raw enrichment (ppm select/non)', 'raw_enr', None, 'Linear FC', None),
                (enr_key, 'bayesian', enr_key, 'log2', fmt_field_bayesian),
                (adjenr_key, 'bayesian', adjenr_key, 'log2', fmt_field_bayesian),
                ('DESeq2 (%s vs Vector)' % bait_label, 'deseq2', 'log2FoldChange_%s_vs_vector_raw' % suffix, 'log2 FC', fmt_field_deseq2),
                (pvec_key, 'bayesian', pvec_key, 'P', fmt_field_bayesian),
                ('DESeq2 (%s vs Vector)' % bait_label, 'deseq2', 'pvalue_%s_vs_vector_raw' % suffix, 'p-value_raw', fmt_field_deseq2),
                ('DESeq2 (%s vs Vector)' % bait_label, 'deseq2', 'pvalue_%s_vs_vector_norm' % suffix, 'p-value_norm', fmt_field_deseq2),
                ('in-frame:Forward', 'junction', ('%sS' % bait_label, 'in_frame'), '%_SEL', fmt_field_select),
                ('in-frame:Forward', 'junction', ('%sN' % bait_label, 'in_frame'), '%_NON', None),
            ]
            for label, kind, extra, unit, fld_fmt in fields:
                ws.write(field_row, col, label, fld_fmt or fmt_field)
                if unit:
                    ws.write(mid_row, col, unit, fmt_unit)
                bait_field_cols.append((bait_label, kind, extra, col))
                col += 1
            block_fmt = fmt_group_bait1 if bait_label == 'Bait1' else fmt_group_bait2
            ws.merge_range(group_row, start_col, group_row, col - 1, '%s stats' % bait_label, block_fmt)

        # Vector stats block. ppm/junction values are averaged across
        # Vector1/Vector2 (2 real replicates) - see vector_mean_ppm() above;
        # junction stats are averaged the same way below. The two Non-Select-
        # only lines (out-of-frame, backwards) are Vector-only by design: they
        # characterize the starting prey pool's own junction quality, which
        # should be shared across Bait1/Bait2's non-selected pools too (same
        # library, no selection pressure applied yet), so one reference
        # column is enough rather than repeating it per bait.
        vector_field_cols = []  # (kind, extra, worksheet_col)
        start_col = col
        vector_fields = [
            ('ppm non', 'ppm_non', None, None, None),
            ('ppm select', 'ppm_sel', None, None, None),
            ('raw enrichment (ppm select/non)', 'raw_enr', None, 'Linear FC', None),
            ('Vector Enrichment (DESeq2)', 'deseq2', 'log2FoldChange_vector', 'log2 FC', fmt_field_deseq2),
            ('Vector Enrichment (DESeq2)', 'deseq2', 'pvalue_vector', 'p-value', fmt_field_deseq2),
            ('in-frame:Forward', 'junction', ('S', 'in_frame'), '%_SEL', fmt_field_select),
            ('in-frame:Forward', 'junction', ('N', 'in_frame'), '%_NON', None),
            ('out-frame:Forward', 'junction', ('N', 'not_in_frame'), '%_NON', None),
            ('backward', 'junction', ('N', 'backwards'), '%_NON', None),
        ]
        for label, kind, extra, unit, fld_fmt in vector_fields:
            ws.write(field_row, col, label, fld_fmt or fmt_field)
            if unit:
                ws.write(mid_row, col, unit, fmt_unit)
            vector_field_cols.append((kind, extra, col))
            col += 1
        ws.merge_range(group_row, start_col, group_row, col - 1, 'Vector stats', fmt_group_vector)

        # ---- Full per-contrast/per-dataset report, appended to the right ----
        # This is the original report layout (DESeq2 blocks in blue, the
        # Bayesian/MCMC block in purple, per-dataset junction breakdown in
        # gray) that predates the Bait1/Bait2/Vector-stats blocks above -
        # kept alongside them, not replaced, so both views are in one report.
        # Block titles for this section sit on mid_row (group_row has no
        # entries here), one row lower than the Bait1/Bait2/Vector titles.
        # Skipped (include_old_report=False) when rebuilding from a loaded
        # general .csv - that source doesn't carry the extra per-dataset
        # junction category breakdown (in_orf/upstream/downstream/etc, only
        # in_frame survives the round-trip) this section needs, and writing
        # it with fabricated zeros would be actively misleading rather than
        # just incomplete.
        stat_field_cols, bayesian_field_cols, junction_field_cols = [], [], []
        if include_old_report:
            deseq2_blocks = [('Bait1 Enrichment (DESeq2)', 'bait1', ['pvalue', 'log2FoldChange', 'Enrichment_score'])]
            if has_bait2:
                deseq2_blocks.append(('Bait2 Enrichment (DESeq2)', 'bait2', ['pvalue', 'log2FoldChange', 'Enrichment_score']))
            deseq2_blocks.append(('Vector Enrichment (DESeq2)', 'vector', ['pvalue', 'log2FoldChange', 'Enrichment_score']))
            if has_bait2:
                deseq2_blocks.append(('Bait1 vs Bait2 Specificity (DESeq2)', 'specificity', ['pvalue', 'log2FoldChange']))
            # suffix here matches the _raw-suffixed keys _suffix_bait_vs_vector
            # renamed Bait-vs-Vector to in `stats` (v7's raw/norm split) -
            # this old-report block duplicates the same raw-counts/poscounts
            # p-value already shown in the main Bait1/Bait2 stats block above,
            # not the newer p-value_norm (which has its own column there).
            deseq2_blocks.append(('Bait1 vs Vector (DESeq2)', 'bait1_vs_vector_raw', ['pvalue', 'log2FoldChange']))
            if has_bait2:
                deseq2_blocks.append(('Bait2 vs Vector (DESeq2)', 'bait2_vs_vector_raw', ['pvalue', 'log2FoldChange']))

            for label, suffix, fields in deseq2_blocks:
                start_col = col
                for fld in fields:
                    stats_key = '%s_%s' % (fld, suffix)
                    ws.write(field_row, col, fld, fmt_field)
                    stat_field_cols.append((stats_key, col))
                    col += 1
                ws.merge_range(mid_row, start_col, mid_row, col - 1, label, fmt_group_deseq2)

            if has_bait2:
                bayesian_fields = ['AdjEnr1', 'AdjEnr2', 'pBait1_Vec', 'pBait2_Vec',
                                   'pBait1_Bait2', 'pBait2_Bait1', 'pBait1', 'pBait2']
            else:
                bayesian_fields = ['AdjEnr', 'p']
            start_col = col
            for fld in bayesian_fields:
                ws.write(field_row, col, fld, fmt_field)
                bayesian_field_cols.append((fld, col))
                col += 1
            ws.merge_range(mid_row, start_col, mid_row, col - 1, 'Bayesian Enrichment (MCMC/JAGS)', fmt_group_bayesian)

            JUNCTION_COLUMNS = ['inframe_inorf', 'upstream', 'in_orf', 'downstream', 'in_frame', 'backwards', 'intron']
            JUNCTION_COLUMN_KEYS = {'inframe_inorf': 'frame_orf'}
            dataset_order = ['Bait1N', 'Bait1S'] + (['Bait2N', 'Bait2S'] if has_bait2 else []) + \
                            ['Vector1N', 'Vector2N', 'Vector1S', 'Vector2S']
            for k in dataset_order:
                start_col = col
                for jc in JUNCTION_COLUMNS:
                    ws.write(field_row, col, jc, fmt_field)
                    junction_field_cols.append((k, jc, col))
                    col += 1
                ws.merge_range(mid_row, start_col, mid_row, col - 1, dataset_labels[k], fmt_group_junction)

        last_col = col - 1

        # Hit-criteria filter/sort: genes passing all of criteria's set
        # thresholds move to the top, sorted descending by sort_key, and
        # get shaded; everyone else follows in the existing gene order.
        highlighted = set()
        gene_order = genes
        if sort_spec is not None and criteria is not None:
            bait_num, sort_key = sort_spec
            maps = _build_bait_value_maps(bait_num, genes, ppm_data, stats, bayesian_stats, junction_fn)
            passing = [g for g in genes if _gene_passes_criteria(g, criteria, maps)]
            passing.sort(key=lambda g: _sort_key_value(g, sort_key, maps), reverse=True)
            highlighted = set(passing)
            remaining = [g for g in genes if g not in highlighted]
            gene_order = passing + remaining
            log_cb("Hit criteria: %d of %d genes passed (Bait%d, sorted by %s)"
                   % (len(passing), len(genes), bait_num, sort_key))

        data_row = field_row + 1
        for gene in gene_order:
            is_hit = gene in highlighted
            num_fmt = fmt_num2_orange if is_hit else fmt_num2
            ws.write(data_row, 0, gene, fmt_gene_orange if is_hit else None)
            srow = stats.get(gene, {})
            brow = bayesian_stats.get(gene, {})

            for bait_label, kind, extra, wcol in bait_field_cols:
                sel_key, non_key = '%sS' % bait_label, '%sN' % bait_label
                if kind == 'ppm_non':
                    ws.write_number(data_row, wcol, ppm_data.get(non_key, {}).get(gene, 0.0), num_fmt)
                elif kind == 'ppm_sel':
                    ws.write_number(data_row, wcol, ppm_data.get(sel_key, {}).get(gene, 0.0), num_fmt)
                elif kind == 'raw_enr':
                    p_sel = ppm_data.get(sel_key, {}).get(gene, 0.0)
                    p_non = ppm_data.get(non_key, {}).get(gene, 0.0)
                    ws.write_number(data_row, wcol, _raw_enrichment_ratio(p_sel, p_non), num_fmt)
                elif kind == 'bayesian':
                    # Leave genuinely blank (not '') when missing - see the
                    # sorting-order note further down for why this matters.
                    if extra in brow and brow[extra] not in ('NA', ''):
                        ws.write_number(data_row, wcol, float(brow[extra]), num_fmt)
                elif kind == 'deseq2':
                    if extra in srow and srow[extra] not in ('NA', ''):
                        ws.write_number(data_row, wcol, float(srow[extra]), num_fmt)
                elif kind == 'junction':
                    dkey, jkey = extra
                    gstat = junction_fn(dkey, gene)
                    ws.write_number(data_row, wcol, gstat[jkey], num_fmt)

            for kind, extra, wcol in vector_field_cols:
                if kind == 'ppm_non':
                    ws.write_number(data_row, wcol, vector_mean_ppm(gene, 'N'), num_fmt)
                elif kind == 'ppm_sel':
                    ws.write_number(data_row, wcol, vector_mean_ppm(gene, 'S'), num_fmt)
                elif kind == 'raw_enr':
                    p_sel = vector_mean_ppm(gene, 'S')
                    p_non = vector_mean_ppm(gene, 'N')
                    ws.write_number(data_row, wcol, _raw_enrichment_ratio(p_sel, p_non), num_fmt)
                elif kind == 'deseq2':
                    if extra in srow and srow[extra] not in ('NA', ''):
                        ws.write_number(data_row, wcol, float(srow[extra]), num_fmt)
                elif kind == 'junction':
                    cond, jkey = extra
                    d1 = junction_fn('Vector1' + cond, gene)
                    d2 = junction_fn('Vector2' + cond, gene)
                    ws.write_number(data_row, wcol, (d1[jkey] + d2[jkey]) / 2, num_fmt)

            for stats_key, wcol in stat_field_cols:
                if stats_key in srow and srow[stats_key] not in ('NA', ''):
                    ws.write_number(data_row, wcol, float(srow[stats_key]), num_fmt)
            for fld, wcol in bayesian_field_cols:
                if fld in brow and brow[fld] not in ('NA', ''):
                    ws.write_number(data_row, wcol, float(brow[fld]), num_fmt)
            for k, jc, wcol in junction_field_cols:
                gstat = junction_fn(k, gene)
                ws.write_number(data_row, wcol, gstat[JUNCTION_COLUMN_KEYS.get(jc, jc)], num_fmt)

            data_row += 1

        # Sort/filter dropdowns on the field-name row, pivot-table style.
        ws.autofilter(field_row, 0, data_row - 1, last_col)
        ws.freeze_panes(data_row - len(genes), 1)
        workbook.close()

    # ---------- hit-criteria panel (v6) ----------

    def _build_export_panel(self):
        """Third column, to the right of the existing file-drop panel -
        built in Python rather than the .ui file since it's dynamic and
        repeated per-bait. The criteria widgets (clickers, sort combo) are
        always interactive - there's no reason to block setting up
        criteria before a run exists. Only the actions that need actual
        results (Export.csv, Find, Export_Sorted_B1/B2) start disabled;
        _collate_finished enables them (and shows/hides the Bait2 section)
        once a run has completed."""
        panel = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(panel)

        self.export_default_btn = QtWidgets.QPushButton('Export.csv')
        self.export_default_btn.setEnabled(False)
        self.export_default_btn.clicked.connect(self.export_default_clicked)
        layout.addWidget(self.export_default_btn)

        layout.addWidget(QtWidgets.QLabel('Statmaker file name (.csv):'))
        self.export_name_edit = QtWidgets.QLineEdit()
        layout.addWidget(self.export_name_edit)
        self.export_data_btn = QtWidgets.QPushButton('Export Data (.csv)')
        self.export_data_btn.setEnabled(False)
        self.export_data_btn.clicked.connect(self.export_data_csv_clicked)
        layout.addWidget(self.export_data_btn)

        load_btn = QtWidgets.QPushButton('Load SM File...')
        load_btn.clicked.connect(self.load_sm_file_clicked)
        layout.addWidget(load_btn)
        self.source_label = QtWidgets.QLabel('Active source: (none yet)')
        self.source_label.setWordWrap(True)
        layout.addWidget(self.source_label)

        self.b1_criteria, b1_group = self._build_criteria_section('Bait1')
        self.b2_criteria, self.bait2_criteria_group = self._build_criteria_section('Bait2')
        layout.addWidget(b1_group)
        layout.addWidget(self.bait2_criteria_group)
        layout.addStretch(1)

        self.export_panel = panel
        self.mainLayout.addWidget(panel)

    def _build_criteria_section(self, bait_label):
        bait_num = 1 if bait_label == 'Bait1' else 2
        group = QtWidgets.QGroupBox('%s Hit Criteria' % bait_label)
        layout = QtWidgets.QVBoxLayout(group)

        def row(label_text, selector):
            r = QtWidgets.QHBoxLayout()
            r.addWidget(QtWidgets.QLabel(label_text))
            r.addWidget(selector)
            layout.addLayout(r)

        p_sel = ClickThroughSelector(
            [None] + [round(0.1 * i, 1) for i in range(1, 10)],
            lambda v: 'none' if v is None else 'P > %.1f' % v)
        pval_raw_sel = ClickThroughSelector(
            [None, 0.1, 0.01, 0.001, 0.0001],
            lambda v: 'none' if v is None else 'p ≤ %g' % v)
        pval_norm_sel = ClickThroughSelector(
            [None, 0.1, 0.01, 0.001, 0.0001],
            lambda v: 'none' if v is None else 'p ≤ %g' % v)
        infr_sel = ClickThroughSelector(
            list(range(0, 101)),
            lambda v: 'none' if v == 0 else '≥ %d%%' % v)
        # Dual enrichment-agreement gate (AdjEnr and DESeq2's own log2FC,
        # both Bait vs Vector, each independently thresholded) replaced the
        # old single raw select/non ratio fold check - see
        # _gene_passes_criteria's docstring for why. Both already log2, so
        # the fold-multiple shown is 2**v, not a linear-ratio conversion.
        fold_values = [None, 0.0, math.log2(1.5), math.log2(2), math.log2(3), math.log2(5)]
        fold_display = lambda v: 'none' if v is None else '> %gx' % round(2 ** v, 4)
        adjenr_fold_sel = ClickThroughSelector(fold_values, fold_display)
        deseq2_fold_sel = ClickThroughSelector(fold_values, fold_display)
        # Defaults, locked in from the rab_hit_recapitulation sweep
        # (2026-08-24) against the prior manually-curated Rab GTPase hit
        # list: infr>=80%, AdjEnr>0, DESeq2>3x, p-value_raw<=0.01 recaptured
        # 95.5% of that list with the least noise of any combination
        # tested. P>0.5 added no further benefit on top of AdjEnr (same
        # posterior, fully redundant there) but P>0.8 does cut further
        # noise, at a real capture cost (95.6%->92.9% in that same sweep) -
        # accepted deliberately. p-value_norm is unreliable per-dataset
        # (see its own label below), so it defaults to off.
        p_sel.index = 8
        p_sel._refresh()
        pval_raw_sel.index = 2
        pval_raw_sel._refresh()
        pval_norm_sel.index = 0
        pval_norm_sel._refresh()
        infr_sel.index = 80
        infr_sel._refresh()
        adjenr_fold_sel.index = 1
        adjenr_fold_sel._refresh()
        deseq2_fold_sel.index = 4
        deseq2_fold_sel._refresh()
        sort_combo = QtWidgets.QComboBox()
        sort_combo.addItems(['DESeq2 (%s vs Vector)' % bait_label, 'Enr%d' % bait_num,
                              'AdjEnr%d' % bait_num, 'ppm/ppm'])

        row('P (probability, MCMC):', p_sel)
        row('p-value_raw (DESeq2):', pval_raw_sel)

        pval_norm_row = QtWidgets.QHBoxLayout()
        pval_norm_label = QtWidgets.QVBoxLayout()
        pval_norm_label.addWidget(QtWidgets.QLabel('p-value_norm (DESeq2):'))
        pval_norm_hint = QtWidgets.QLabel('(good for low crush)')
        hint_font = pval_norm_hint.font()
        hint_font.setPointSize(max(hint_font.pointSize() - 2, 6))
        pval_norm_hint.setFont(hint_font)
        pval_norm_label.addWidget(pval_norm_hint)
        pval_norm_row.addLayout(pval_norm_label)
        pval_norm_row.addWidget(pval_norm_sel)
        layout.addLayout(pval_norm_row)

        row('% in-frame:Forward:', infr_sel)
        row('AdjEnr (%s vs Vector):' % bait_label, adjenr_fold_sel)
        row('DESeq2 Fold (%s vs Vector):' % bait_label, deseq2_fold_sel)
        row('Sort by:', sort_combo)

        find_row = QtWidgets.QHBoxLayout()
        find_btn = QtWidgets.QPushButton('Find')
        count_label = QtWidgets.QLabel('')
        count_font = count_label.font()
        count_font.setPointSize(count_font.pointSize() + 6)
        count_font.setBold(True)
        count_label.setFont(count_font)
        find_row.addWidget(find_btn)
        find_row.addWidget(count_label)
        find_row.addStretch(1)
        layout.addLayout(find_row)

        find_btn.setEnabled(False)
        export_btn = QtWidgets.QPushButton('Export_Sorted_B%d' % bait_num)
        export_btn.setEnabled(False)
        layout.addWidget(export_btn)

        widgets = {
            'p': p_sel, 'pval_raw': pval_raw_sel, 'pval_norm': pval_norm_sel,
            'infr': infr_sel, 'adjenr_fold': adjenr_fold_sel, 'deseq2_fold': deseq2_fold_sel,
            'sort_combo': sort_combo, 'find_btn': find_btn, 'count_label': count_label,
            'export_btn': export_btn,
        }
        if bait_num == 1:
            self.b1_count_label = count_label
            find_btn.clicked.connect(self.find_b1_clicked)
            export_btn.clicked.connect(self.export_sorted_b1_clicked)
        else:
            self.b2_count_label = count_label
            find_btn.clicked.connect(self.find_b2_clicked)
            export_btn.clicked.connect(self.export_sorted_b2_clicked)
        return widgets, group

    RESULTS_COLUMNS = ['Gene', 'ppm non', 'ppm select', 'Enr ppm SEL/NON',
                       'in-frame:Forward (Select)']
    RESULTS_COLUMN_WIDTHS = [80, 70, 70, 120, 100]

    def _build_results_panel(self):
        """Fifth column: per-bait table of the genes currently passing that
        bait's hit criteria (the same list _matching_genes() already
        computes for the Find count) - a quick-scan view, not the full
        report, so only a small fixed subset of columns is shown."""
        panel = QtWidgets.QWidget()
        # Proportionally wider from the start, rather than relying on the
        # table columns' own fixed-width sum (440px) plus whatever padding
        # happens to fall out of the layout - gives the hit-match tables
        # real breathing room next to the criteria panel immediately, no
        # manual resize needed to see the full columns.
        panel.setMinimumWidth(560)
        layout = QtWidgets.QVBoxLayout(panel)

        self.b1_results_table, b1_results_group = self._build_results_table('Bait1')
        self.b2_results_table, self.bait2_results_group = self._build_results_table('Bait2')
        layout.addWidget(b1_results_group)
        layout.addWidget(self.bait2_results_group)

        self.results_panel = panel
        self.mainLayout.addWidget(panel)

    def _build_results_table(self, bait_label):
        group = QtWidgets.QGroupBox('%s Matches' % bait_label)
        layout = QtWidgets.QVBoxLayout(group)
        table = QtWidgets.QTableWidget(0, len(self.RESULTS_COLUMNS))
        table.setHorizontalHeaderLabels(self.RESULTS_COLUMNS)
        table.horizontalHeader().setDefaultAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        table.verticalHeader().setVisible(False)
        table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        for col, width in enumerate(self.RESULTS_COLUMN_WIDTHS):
            table.setColumnWidth(col, width)
        layout.addWidget(table)
        return table, group

    def _populate_results_table(self, table, bait_num, widgets):
        source = self._active_source
        maps = source['maps1'] if bait_num == 1 else source['maps2']
        sort_key = self._sort_key_from_combo(widgets['sort_combo'])
        genes = sorted(self._matching_genes(bait_num, widgets),
                        key=lambda g: _sort_key_value(g, sort_key, maps), reverse=True)
        ppm_sel = source['ppm_data'].get('Bait%dS' % bait_num, {})
        ppm_non = source['ppm_data'].get('Bait%dN' % bait_num, {})

        table.setRowCount(len(genes))
        for row, g in enumerate(genes):
            values = [g, ppm_non.get(g, 0.0), ppm_sel.get(g, 0.0),
                      maps['ratio'].get(g), maps['infr'].get(g)]
            for col, val in enumerate(values):
                text = val if isinstance(val, str) else ('' if val is None else '%.4g' % val)
                table.setItem(row, col, QtWidgets.QTableWidgetItem(text))

    @staticmethod
    def _sort_key_from_combo(combo):
        return ['deseq2', 'enr', 'adjenr', 'ratio'][combo.currentIndex()]

    def _criteria_dict(self, widgets):
        return {'p': widgets['p'].value, 'pval_raw': widgets['pval_raw'].value,
               'pval_norm': widgets['pval_norm'].value, 'infr': widgets['infr'].value,
               'adjenr_fold': widgets['adjenr_fold'].value, 'deseq2_fold': widgets['deseq2_fold'].value}

    def _matching_genes(self, bait_num, widgets):
        source = self._active_source
        maps = source['maps1'] if bait_num == 1 else source['maps2']
        criteria = self._criteria_dict(widgets)
        return [g for g in source['genes'] if _gene_passes_criteria(g, criteria, maps)]

    @staticmethod
    def _n_above_3ppm_selected(source, bait_num):
        """Denominator for the Find count: how many genes are even still
        above 3ppm (the same detection floor used elsewhere) in that
        bait's own Selected sample - i.e. how many candidates survived
        selection at all, before any hit criteria are applied."""
        sel = source['ppm_data'].get('Bait%dS' % bait_num, {})
        return sum(1 for g in source['genes'] if sel.get(g, 0.0) >= 3.0)

    @QtCore.pyqtSlot()
    def find_b1_clicked(self):
        n = len(self._matching_genes(1, self.b1_criteria))
        d = self._n_above_3ppm_selected(self._active_source, 1)
        self.b1_count_label.setText("%d / %d" % (n, d))
        self._populate_results_table(self.b1_results_table, 1, self.b1_criteria)

    @QtCore.pyqtSlot()
    def find_b2_clicked(self):
        n = len(self._matching_genes(2, self.b2_criteria))
        d = self._n_above_3ppm_selected(self._active_source, 2)
        self.b2_count_label.setText("%d / %d" % (n, d))
        self._populate_results_table(self.b2_results_table, 2, self.b2_criteria)

    def _source_write_args(self, source):
        """Common args _write_workbook needs, regardless of whether
        `source` came from a fresh run (self._last_run, real bqp_data) or
        a loaded general .csv (_parse_sm_csv, flattened junction_fn).
        include_old_report is False for a loaded file - see _parse_sm_csv's
        docstring for what doesn't survive that round-trip."""
        if source['kind'] == 'run':
            run = self._last_run
            return dict(csv_paths=run['csv_paths'], dataset_labels=run['dataset_labels'],
                       overdispersion=run['overdispersion'], bayes_overdispersion=run['bayes_overdispersion'],
                       stats=run['stats'], bayesian_stats=run['bayesian_stats'], bqp_data=run['bqp_data'],
                       ppm_data=run['ppm_data'], junction_fn=_default_junction_fn(run['bqp_data']),
                       include_old_report=True)
        return dict(csv_paths=source['csv_paths'], dataset_labels=source['dataset_labels'],
                   overdispersion=source['overdispersion'], bayes_overdispersion=source['bayes_overdispersion'],
                   stats=source['stats'], bayesian_stats=source['bayesian_stats'], bqp_data={},
                   ppm_data=source['ppm_data'], junction_fn=source['junction_fn'],
                   include_old_report=False)

    @QtCore.pyqtSlot()
    def export_default_clicked(self):
        source = self._active_source
        args = self._source_write_args(source)
        timestamp = datetime.now().strftime('%d%b%Y-%H%M%S')
        out_name = 'stat_maker_%s.xlsx' % timestamp
        out_path = os.path.join(source.get('out_folder') or os.path.dirname(source.get('src_path', '.')), out_name)
        self._write_workbook(out_path, args['csv_paths'], args['dataset_labels'], args['overdispersion'],
                             args['bayes_overdispersion'], args['stats'], args['bayesian_stats'], args['bqp_data'],
                             source['genes'], source['has_bait2'], self.log,
                             ppm_data=args['ppm_data'], junction_fn=args['junction_fn'],
                             include_old_report=args['include_old_report'])
        self.log("Wrote %s" % out_path)
        self.statusbar.showMessage("Wrote %s" % out_name, 8000)

    def _export_sorted(self, bait_num, widgets, filename_prefix):
        source = self._active_source
        sort_key = self._sort_key_from_combo(widgets['sort_combo'])
        criteria = self._criteria_dict(widgets)
        args = self._source_write_args(source)
        timestamp = datetime.now().strftime('%d%b%Y-%H%M%S')
        out_name = '%sstat_maker_%s.xlsx' % (filename_prefix, timestamp)
        out_path = os.path.join(source.get('out_folder') or os.path.dirname(source.get('src_path', '.')), out_name)
        self._write_workbook(out_path, args['csv_paths'], args['dataset_labels'], args['overdispersion'],
                             args['bayes_overdispersion'], args['stats'], args['bayesian_stats'], args['bqp_data'],
                             source['genes'], source['has_bait2'], self.log,
                             sort_spec=(bait_num, sort_key), criteria=criteria,
                             ppm_data=args['ppm_data'], junction_fn=args['junction_fn'],
                             include_old_report=args['include_old_report'])
        self.log("Wrote %s" % out_path)
        self.statusbar.showMessage("Wrote %s" % out_name, 8000)

    @QtCore.pyqtSlot()
    def export_sorted_b1_clicked(self):
        self._export_sorted(1, self.b1_criteria, 'SORT1_')

    @QtCore.pyqtSlot()
    def export_sorted_b2_clicked(self):
        self._export_sorted(2, self.b2_criteria, 'SORT2_')

    @staticmethod
    def _xlsx_to_csv(xlsx_path, csv_path):
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        with open(csv_path, 'w', newline='') as f:
            w = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                vals = list(row)
                while vals and vals[-1] is None:
                    vals.pop()
                w.writerow(["" if v is None else v for v in vals])

    @QtCore.pyqtSlot()
    def export_data_csv_clicked(self):
        """The general .csv - a flat dump of the exact same rich report
        the .xlsx exports produce (see _parse_sm_csv's docstring). Built
        by writing to a temp .xlsx and converting it, then discarding the
        temp file - this is what Load SM File... reads back later."""
        run = self._last_run
        if run is None:
            return
        name = self.export_name_edit.text().strip() or datetime.now().strftime('stat_maker_%d%b%Y-%H%M%S')
        out_path = os.path.join(run['out_folder'], name + '.csv')
        self.log("Building report (temp xlsx) for %s ..." % out_path)
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_xlsx = tmp.name
        try:
            self._write_workbook(tmp_xlsx, run['csv_paths'], run['dataset_labels'], run['overdispersion'],
                                 run['bayes_overdispersion'], run['stats'], run['bayesian_stats'], run['bqp_data'],
                                 run['genes'], run['has_bait2'], self.log, ppm_data=run['ppm_data'])
            self.log("Converting to %s ..." % out_path)
            self._xlsx_to_csv(tmp_xlsx, out_path)
        finally:
            os.remove(tmp_xlsx)
        self.log("Wrote %s" % out_path)
        self.statusbar.showMessage("Wrote %s" % os.path.basename(out_path), 8000)

    @QtCore.pyqtSlot()
    def load_sm_file_clicked(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Load Stat Maker Data", self.directory, "CSV files (*.csv)")
        if not path:
            return
        try:
            parsed = _parse_sm_csv(path)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Load Failed", str(e))
            return
        parsed['kind'] = 'file'
        parsed['out_folder'] = os.path.dirname(path)
        parsed['maps1'] = _build_bait_value_maps(1, parsed['genes'], parsed['ppm_data'], parsed['stats'],
                                                  parsed['bayesian_stats'], parsed['junction_fn'])
        parsed['maps2'] = (_build_bait_value_maps(2, parsed['genes'], parsed['ppm_data'], parsed['stats'],
                                                   parsed['bayesian_stats'], parsed['junction_fn'])
                          if parsed['has_bait2'] else None)
        self._active_source = parsed
        self.export_default_btn.setEnabled(True)
        self.b1_criteria['find_btn'].setEnabled(True)
        self.b1_criteria['export_btn'].setEnabled(True)
        self.b2_criteria['find_btn'].setEnabled(parsed['has_bait2'])
        self.b2_criteria['export_btn'].setEnabled(parsed['has_bait2'])
        self.bait2_criteria_group.setVisible(parsed['has_bait2'])
        self.bait2_results_group.setVisible(parsed['has_bait2'])
        self.b1_count_label.setText('')
        self.b2_count_label.setText('')
        self.b1_results_table.setRowCount(0)
        self.b2_results_table.setRowCount(0)
        self.source_label.setText("Active source: %s" % parsed['name'])
        self.export_name_edit.setText(parsed['name'])
        self.log("Loaded %s (%d genes%s)"
                 % (path, len(parsed['genes']), ", Bait2 present" if parsed['has_bait2'] else ""))

    @QtCore.pyqtSlot()
    def method_info_clicked(self):
        try:
            with open(R_SCRIPT_PATH, encoding='utf-8') as fh:
                r_source = fh.read()
        except (OSError, UnicodeDecodeError) as e:
            r_source = "(Could not read %s: %s)" % (R_SCRIPT_PATH, e)

        dialog = QtWidgets.QDialog(self)
        dialog.setWindowTitle("Method Info")
        dialog.resize(800, 700)
        layout = QtWidgets.QVBoxLayout(dialog)
        text = QtWidgets.QPlainTextEdit(dialog)
        text.setReadOnly(True)
        text.setFont(QtGui.QFont("Menlo", 12))
        text.setPlainText(METHOD_INFO_TEXT + r_source)
        layout.addWidget(text)
        close_btn = QtWidgets.QPushButton("Close", dialog)
        close_btn.clicked.connect(dialog.accept)
        layout.addWidget(close_btn)
        dialog.exec_()

    @QtCore.pyqtSlot()
    def quit_clicked(self):
        app.quit()
        sys.exit()


def appExit():
    app.quit()
    sys.exit()

if __name__ == '__main__':
    form = Stat_Maker_Gui()
    form.show()
    app.aboutToQuit.connect(appExit)
    app.exec_()
