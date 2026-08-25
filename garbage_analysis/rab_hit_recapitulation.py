#!/usr/bin/env python3
"""Finds hit-criteria thresholds that recapitulate the prior manually-curated
Rab GTPase interactor list (TAB_RRAB_Hits.xlsx) from the new Stat Maker v7
output (36 SM6_<Rab>_23Aug2026.csv files, DESeq2 raw+norm p-values +
Bayesian/MCMC). Staged/greedy sweep, priority order per user instruction:
in-frame:Forward -> Enrichment Fold -> P (MCMC) -> p-value_raw -> p-value_norm.
"""
import sys, os, re, csv, zipfile, math
from openpyxl.utils import column_index_from_string

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, '/Users/robertpiper2/LOCAL_DEEPN')
os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
os.chdir('/Users/robertpiper2/LOCAL_DEEPN')

import openpyxl
import stat_maker_gui_v7 as smg
import sm6_batch as batch
import functions.stat_collation as sc

TAB_PATH = '/Users/robertpiper2/endosomeLAB Dropbox/Rob Piper/SM6_batch_output/TAB_RRAB_Hits.xlsx'
GARBAGE_DOCX = '/Users/robertpiper2/endosomeLAB Dropbox/Rob Piper/SM6_batch_output/HuORF Garbage.docx'
OUT_CSV = '/Users/robertpiper2/endosomeLAB Dropbox/Rob Piper/SM6_batch_output/Rab_hit_criteria_sweep.csv'

# Historical spreadsheet's Rab-column prefix -> canonical rid used by SM6 files.
# Verified by cross-checking every canonical dataset's own Bait1/Bait2 filenames
# (Bait1 always contains "QL", Bait2 always contains "TN" - a handful of Rabs
# in the historical sheet label that same Bait2 construct "_SN" instead of
# "_TN", an old naming variant for the identical construct, not a different one).
HIST_TO_RID = {
    'Rab5A': 'Rab5A', 'Rab4A': 'Rab4A', 'Rab14': 'Rab14', 'Rab6A': 'Rab6A',
    'Rab8A': 'Rab8', 'Rab9A': 'Rab9', 'Rab27B': 'Rab27B', 'Rab12': 'Rab12',
    'Rab13': 'Rab13', 'Rab15': 'Rab15', 'Rab27A': 'Rab27A', 'Rab10': 'Rab10',
    'Rab7A': 'Rab7A', 'Rab2A': 'Rab2', 'Rab21': 'Rab21', 'Rab22A': 'Rab22',
    'Rab23': 'Rab23', 'Rab26': 'Rab26', 'Rab28': 'Rab28', 'Rab33A': 'Rab33A',
    'Rab34': 'Rab34', 'Rab35': 'Rab35', 'Rab36': 'Rab36', 'Rab37': 'Rab37',
    'Rab38': 'Rab38', 'Rab3A': 'Rab3A', 'Rab39A': 'Rab39A', 'Rab43': 'Rab43',
    'Rab11A': 'Rab11', 'Rab17': 'Rab17', 'Rab19': 'Rab19', 'Rab30': 'Rab30',
    'Rab31': 'Rab31', 'Rab32': 'Rab32', 'Rab1A*': 'Rab1star', 'Rab18*': 'Rab18star',
}

# Extracted directly from HuORF Garbage.docx (verified via a regex parse of
# the gene/count/X table, 30 genes, cross-checked against the raw text).
GARBAGE_GENES = {
    'CYB5R4', 'HSPA8', 'EFNB1', 'SEZ6L', 'ACLY', 'MPDZ', 'STAU2', 'RNF40',
    'CDC42EP1', 'SUGT1', 'KRT222', 'RLIM', 'GOPC', 'CRX', 'PAX7', 'RNF2',
    'POLK', 'ZNF846', 'VCAM1', 'TXNL1', 'TBCCD1', 'BCORL1', 'PCHDGB4',
    'ZNF561', 'ERGIC3', 'RWDD1', 'LRRN1', 'EIF2S2', 'POLR2C', 'C10orf2',
}


def load_reference_hits():
    wb = openpyxl.load_workbook(TAB_PATH, data_only=True)
    ws = wb['Sheet1']
    bx = column_index_from_string('BX')
    es = column_index_from_string('ES')
    header = [ws.cell(row=2, column=c).value for c in range(bx, es + 1)]

    col_meta = []  # (col_idx, rid, bait_num)
    unmapped = set()
    for i, label in enumerate(header):
        if not label:
            continue
        m = re.match(r'^(.*)_(QL|TN|SN)$', label)
        if not m:
            continue
        prefix, suffix = m.groups()
        rid = HIST_TO_RID.get(prefix)
        if rid is None:
            unmapped.add(prefix)
            continue
        bait_num = 1 if suffix == 'QL' else 2
        col_meta.append((bx + i, rid, bait_num))
    if unmapped:
        print("WARNING: unmapped historical-sheet prefixes:", unmapped)

    ref = {}
    for r in range(3, ws.max_row + 1):
        gene = ws.cell(row=r, column=bx).value
        if not gene:
            continue
        for col_idx, rid, bait_num in col_meta:
            if ws.cell(row=r, column=col_idx).value == 'HIT':
                ref.setdefault(rid, {}).setdefault(bait_num, set()).add(gene)

    total = sum(len(s) for baits in ref.values() for s in baits.values())
    print("Reference hit sets: %d Rab-bait entries covering %d Rabs, %d total HIT calls"
          % (len(col_meta), len(ref), total))
    return ref, col_meta


def load_all_sm6_data():
    data = {}
    for fn in batch.CANONICAL:
        rid = batch.rab_id(fn)
        path = os.path.join(batch.OUT_DIR, 'SM6_%s_23Aug2026.csv' % rid)
        parsed = smg._parse_sm_csv(path)
        maps1 = smg._build_bait_value_maps(1, parsed['genes'], parsed['ppm_data'], parsed['stats'],
                                            parsed['bayesian_stats'], parsed['junction_fn'])
        maps2 = (smg._build_bait_value_maps(2, parsed['genes'], parsed['ppm_data'], parsed['stats'],
                                             parsed['bayesian_stats'], parsed['junction_fn'])
                 if parsed['has_bait2'] else None)
        data[rid] = {'genes': parsed['genes'], 1: maps1, 2: maps2}
        print("  loaded", rid, len(parsed['genes']), "genes")
    return data


def gene_passes_criteria_v2(gene, criteria, maps):
    """Like smg._gene_passes_criteria, except the enrichment-fold check
    reads criteria['enrichment_source'] ('deseq2' = DESeq2 Bait-vs-Vector
    log2FoldChange, or 'adjenr' = Bayesian/MCMC posterior median log2
    effect size, Bait vs Vector) instead of the raw within-bait select/non
    ratio - both of those are already log2-scale, so criteria['fold'] is a
    log2 threshold here (compared directly), not a linear multiple run
    through _signed_fold."""
    p = criteria.get('p')
    if p is not None:
        val = maps['p'].get(gene)
        if val is None or not (val > p):
            return False

    pval_raw = criteria.get('pval_raw')
    if pval_raw is not None:
        val = maps['pval_raw'].get(gene)
        if val is None or not (val <= pval_raw):
            return False

    pval_norm = criteria.get('pval_norm')
    if pval_norm is not None:
        val = maps['pval_norm'].get(gene)
        if val is None or not (val <= pval_norm):
            return False

    infr = criteria.get('infr') or 0
    if infr:
        val = maps['infr'].get(gene, 0)
        if not (val >= infr):
            return False

    fold = criteria.get('fold')
    if fold is not None:
        source = criteria.get('enrichment_source', 'deseq2')
        val = maps[source].get(gene)
        if val is None or not (val > fold):
            return False

    return True


def evaluate(criteria, ref, col_meta, data, passes_fn=None):
    """Runs `criteria` against every (rid, bait_num) in col_meta. Returns
    aggregate stats plus a per-(rid,bait_num) breakdown dict."""
    passes_fn = passes_fn or smg._gene_passes_criteria
    total_ref = 0
    total_captured = 0
    total_extra = 0
    total_garbage_hit = 0
    per_dataset = {}
    for _col, rid, bait_num in col_meta:
        maps = data[rid][bait_num]
        if maps is None:
            continue
        genes = data[rid]['genes']
        ref_set = ref.get(rid, {}).get(bait_num, set())
        passing = {g for g in genes if passes_fn(g, criteria, maps)}
        captured = passing & ref_set
        garbage_hit = passing & GARBAGE_GENES
        extra = passing - ref_set - GARBAGE_GENES
        total_ref += len(ref_set)
        total_captured += len(captured)
        total_extra += len(extra)
        total_garbage_hit += len(garbage_hit)
        per_dataset[(rid, bait_num)] = {
            'ref': len(ref_set), 'captured': len(captured),
            'extra': len(extra), 'garbage_hit': len(garbage_hit),
        }
    capture_pct = (100.0 * total_captured / total_ref) if total_ref else 0.0
    return {
        'capture_pct': capture_pct, 'total_ref': total_ref, 'total_captured': total_captured,
        'total_extra': total_extra, 'total_garbage_hit': total_garbage_hit,
        'per_dataset': per_dataset,
    }


def make_row(label, criteria, result, col_meta):
    row = {
        'step': label,
        'P': criteria.get('p'), 'pval_raw': criteria.get('pval_raw'),
        'pval_norm': criteria.get('pval_norm'), 'infr': criteria.get('infr'),
        'fold': criteria.get('fold'), 'enrichment_source': criteria.get('enrichment_source', ''),
        'total_pct_captured': round(result['capture_pct'], 2),
        'total_ref_hits': result['total_ref'],
        'total_captured': result['total_captured'],
        'total_extra_hits': result['total_extra'],
        'total_garbage_excluded': result['total_garbage_hit'],
    }
    for _col, rid, bait_num in col_meta:
        d = result['per_dataset'].get((rid, bait_num))
        key = '%s_Bait%d' % (rid, bait_num)
        row[key] = ('%d/%d' % (d['captured'], d['ref'])) if d else ''
    return row


def main():
    print("=== Loading reference hit list ===")
    ref, col_meta = load_reference_hits()

    print("=== Loading garbage gene list ===")
    print("  %d garbage genes: %s" % (len(GARBAGE_GENES), sorted(GARBAGE_GENES)))

    print("=== Loading all 36 SM6 datasets ===")
    data = load_all_sm6_data()

    rows = []
    fieldnames = ['step', 'P', 'pval_raw', 'pval_norm', 'infr', 'fold', 'enrichment_source',
                  'total_pct_captured', 'total_ref_hits', 'total_captured',
                  'total_extra_hits', 'total_garbage_excluded']
    fieldnames += ['%s_Bait%d' % (rid, bait_num) for _c, rid, bait_num in col_meta]

    # ---- Stage 1: in-frame:Forward x Enrichment source x fold grid ----
    # enrichment_source: 'deseq2' (DESeq2 Bait-vs-Vector log2FC) or 'adjenr'
    # (Bayesian/MCMC posterior median log2 effect size, Bait vs Vector) -
    # both log2-scale, unlike the raw within-bait select/non ratio, which
    # structurally can't recover "low-crush" hits (de-enriched within the
    # bait's own culture, but still more enriched than Vector).
    print("=== Stage 1: in-frame x enrichment-source x fold grid ===")
    infr_candidates = [50, 80, 90]
    fold_multiples = [1.0, 1.5, 2, 3, 5]  # 1.0x -> log2 threshold 0 (any positive log2FC)
    sources = ['deseq2', 'adjenr']
    stage1_results = []
    for source in sources:
        for infr in infr_candidates:
            for fold_x in fold_multiples:
                fold_log2 = math.log2(fold_x)
                criteria = {'p': None, 'pval_raw': None, 'pval_norm': None, 'infr': infr,
                            'fold': fold_log2, 'enrichment_source': source}
                result = evaluate(criteria, ref, col_meta, data, gene_passes_criteria_v2)
                label = '%s: infr>=%d%%, fold>%glog2(%gx)' % (source, infr, fold_log2, fold_x)
                print("  %-38s capture=%.1f%% extra=%d" % (label, result['capture_pct'], result['total_extra']))
                rows.append(make_row(label, criteria, result, col_meta))
                stage1_results.append((source, infr, fold_x, criteria, result))

    # Pick the tightest combo that still clears 95% aggregate capture -
    # "more than 95 favored" - tie-broken by minimum extra-hit count.
    passing95 = [t for t in stage1_results if t[4]['capture_pct'] >= 95.0]
    if passing95:
        best = min(passing95, key=lambda t: t[4]['total_extra'])
    else:
        best = max(stage1_results, key=lambda t: t[4]['capture_pct'])
        print("  WARNING: no combo reached 95%% capture - "
              "using the best available (%.1f%%) as the stage-1 baseline." % best[4]['capture_pct'])
    base_source, base_infr, base_fold_x, base_criteria, base_result = best
    print("=== Stage 1 winner: %s, infr>=%d%%, fold>%gx (capture=%.1f%%, extra=%d) ==="
          % (base_source, base_infr, base_fold_x, base_result['capture_pct'], base_result['total_extra']))
    rows.append(make_row('BASELINE: %s, infr>=%d%%, fold>%gx' % (base_source, base_infr, base_fold_x),
                          base_criteria, base_result, col_meta))

    # ---- Stage 2: P sweep on top of the baseline ----
    print("=== Stage 2: P (MCMC) sweep on baseline ===")
    for p in [0.1, 0.2, 0.3, 0.4, 0.5]:
        criteria = dict(base_criteria, p=p)
        result = evaluate(criteria, ref, col_meta, data, gene_passes_criteria_v2)
        label = 'baseline + P>%.1f' % p
        print("  %-28s capture=%.1f%% extra=%d" % (label, result['capture_pct'], result['total_extra']))
        rows.append(make_row(label, criteria, result, col_meta))

    # ---- Stage 3: p-value_raw sweep on top of the baseline ----
    print("=== Stage 3: p-value_raw sweep on baseline ===")
    for pv in [0.1, 0.01]:
        criteria = dict(base_criteria, pval_raw=pv)
        result = evaluate(criteria, ref, col_meta, data, gene_passes_criteria_v2)
        label = 'baseline + p_raw<=%g' % pv
        print("  %-28s capture=%.1f%% extra=%d" % (label, result['capture_pct'], result['total_extra']))
        rows.append(make_row(label, criteria, result, col_meta))

    # ---- Stage 4: p-value_norm sweep on top of the baseline ----
    print("=== Stage 4: p-value_norm sweep on baseline ===")
    for pv in [0.1, 0.01]:
        criteria = dict(base_criteria, pval_norm=pv)
        result = evaluate(criteria, ref, col_meta, data, gene_passes_criteria_v2)
        label = 'baseline + p_norm<=%g' % pv
        print("  %-28s capture=%.1f%% extra=%d" % (label, result['capture_pct'], result['total_extra']))
        rows.append(make_row(label, criteria, result, col_meta))

    with open(OUT_CSV, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in rows:
            w.writerow(row)
    print("=== Wrote %s (%d rows) ===" % (OUT_CSV, len(rows)))


if __name__ == '__main__':
    main()
