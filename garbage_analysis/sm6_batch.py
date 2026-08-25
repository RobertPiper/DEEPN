#!/usr/bin/env python3
"""Overnight SM6 re-analysis batch: runs the full Stat Maker v6 pipeline
(DESeq2 poscounts + Bayesian/MCMC + junction collation) on all 36 canonical
historical Rab GTPase 3-way datasets, writing SM6_<Rab>_<ddMonYYYY>.csv - a
flat CSV dump of the exact same rich report the app's .xlsx export
produces (metadata, OVERDISPERSION, CRUSH table, three-row header, every
column block) - into Dropbox. No .xlsx is kept; it's built to a temp file,
converted to CSV, then discarded. One dataset failing does not stop the
rest - every outcome is logged and summarized at the end."""
import sys, os, re, time, traceback
from datetime import datetime

os.environ.setdefault('QT_QPA_PLATFORM', 'offscreen')
sys.path.insert(0, '/Users/robertpiper2/LOCAL_DEEPN')
os.chdir('/Users/robertpiper2/LOCAL_DEEPN')

import openpyxl
import csv as csv_mod
import stat_maker_gui_v7 as smg
import functions.stat_collation as sc

SM = "/Volumes/PiperLabDataDisk/DEEPN_2018_RabGTPase/stat_maker_output"
GC = "/Volumes/PiperLabDataDisk/DEEPN_2018_RabGTPase/gene_count_summary/"
BQP_ROOT = "/Volumes/PiperLabDataDisk/DEEPN_2018_RabGTPase"
OUT_DIR = "/Users/robertpiper2/Dropbox (endosomeLAB)/SM6_batch_output"
INTERMEDIATE_ROOT = "/private/tmp/claude-503/-Users-robertpiper2-Desktop/d0bd16a2-ab29-49c2-9981-9811e0c1149f/scratchpad/sm6_batch_intermediate"
LOG_PATH = os.path.join(OUT_DIR, "_batch_log.txt")

CANONICAL = [
  "Rab10_stat_maker_03Jul2018-130449.csv", "Rab11_stat_maker_25Sep2018-081958.csv",
  "Rab12_stat_maker_03Jul2018-175734.csv", "Rab13_stat_maker_03Jul2018-181356.csv",
  "Rab14_stat_maker_03Jul2018-131936.csv", "Rab15_stat_maker_03Jul2018-184306.csv",
  "Rab17_stat_maker_25Sep2018-085912.csv", "Rab18star_stat_maker_21Sep2018-164240.csv",
  "Rab19_stat_maker_25Sep2018-091531.csv", "Rab1star_stat_maker_21Sep2018-165558.csv",
  "Rab21_stat_maker_21Sep2018-112556.csv", "Rab22_stat_maker_21Sep2018-114418.csv",
  "Rab23_stat_maker_21Sep2018-120435.csv", "Rab26_stat_maker_21Sep2018-122236.csv",
  "Rab27A_stat_maker_03Jul2018-173057.csv", "Rab27B_stat_maker_03Jul2018-174313.csv",
  "Rab28_stat_maker_21Sep2018-124022.csv", "Rab2_stat_maker_20Sep2018-165603.csv",
  "Rab30_stat_maker_25Sep2018-093405.csv", "Rab31_stat_maker_25Sep2018-094742.csv",
  "Rab32_stat_maker_25Sep2018-100258.csv", "Rab33A_stat_maker_21Sep2018-134520.csv",
  "Rab34_stat_maker_21Sep2018-140602.csv", "Rab35_stat_maker_21Sep2018-142301.csv",
  "Rab36_stat_maker_24Sep2018-143516.csv", "Rab37_stat_maker_21Sep2018-150805.csv",
  "Rab38_stat_maker_21Sep2018-152515.csv", "Rab39A_stat_maker_21Sep2018-160143.csv",
  "Rab3A_stat_maker_21Sep2018-154418.csv", "Rab43_stat_maker_21Sep2018-162519.csv",
  "Rab4A_stat_maker_03Jul2018-133430.csv", "Rab5A_stat_maker_03Jul2018-134802.csv",
  "Rab6A_stat_maker_03Jul2018-141035.csv", "Rab7A_stat_maker_03Jul2018-142348.csv",
  "Rab8_stat_maker_03Jul2018-170444.csv", "Rab9_stat_maker_03Jul2018-171533.csv",
]

DATASET_LABELS = {
    'Bait1N': 'Bait1_Non-Selected', 'Bait1S': 'Bait1_Selected',
    'Bait2N': 'Bait2_Non-Selected', 'Bait2S': 'Bait2_Selected',
    'Vector1N': 'Vector_Non-Selected_1', 'Vector2N': 'Vector_Non-Selected_2',
    'Vector1S': 'Vector_Selected_1', 'Vector2S': 'Vector_Selected_2',
}


def log(msg):
    line = "[%s] %s" % (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), msg)
    print(line, flush=True)
    with open(LOG_PATH, 'a') as f:
        f.write(line + '\n')


def get_config(path):
    cfg = {}
    with open(path, encoding='latin-1', errors='ignore') as f:
        content = f.read(20000)
    for key in ('Bait1_Non-Selected_1', 'Bait1_Selected_1', 'Bait2_Non-Selected_1', 'Bait2_Selected_1',
                'Vector_Non-Selected_1', 'Vector_Selected_1', 'Vector_Non-Selected_2', 'Vector_Selected_2'):
        m = re.search(re.escape(key) + r'\s*,\s*([^\r\n,]+\.csv)', content)
        cfg[key] = os.path.basename(m.group(1).strip()) if m else None
    return cfg


def rab_id(fname):
    m = re.match(r'^(Rab\w+?)_stat_maker', fname)
    return m.group(1) if m else fname


def xlsx_to_csv(xlsx_path, csv_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    with open(csv_path, 'w', newline='') as f:
        w = csv_mod.writer(f)
        for row in ws.iter_rows(values_only=True):
            vals = list(row)
            while vals and vals[-1] is None:
                vals.pop()
            w.writerow(["" if v is None else v for v in vals])


def run_one(fn, rscript_exe):
    rid = rab_id(fn)
    cfg = get_config(os.path.join(SM, fn))
    csv_paths = {
        'Bait1S': GC + cfg['Bait1_Selected_1'], 'Bait1N': GC + cfg['Bait1_Non-Selected_1'],
        'Bait2S': GC + cfg['Bait2_Selected_1'], 'Bait2N': GC + cfg['Bait2_Non-Selected_1'],
        'Vector1S': GC + cfg['Vector_Selected_1'], 'Vector1N': GC + cfg['Vector_Non-Selected_1'],
        'Vector2S': GC + cfg['Vector_Selected_2'], 'Vector2N': GC + cfg['Vector_Non-Selected_2'],
    }
    for k, p in csv_paths.items():
        if not os.path.exists(p):
            raise FileNotFoundError("%s: %s" % (k, p))
    has_bait2 = True

    run_out_dir = os.path.join(INTERMEDIATE_ROOT, rid)
    os.makedirs(run_out_dir, exist_ok=True)

    log("  [%s] loading raw counts..." % rid)
    raw_count_data = {k: sc.load_raw_counts_summed(p) for k, p in csv_paths.items()}
    collated_path = os.path.join(run_out_dir, 'collated_input.csv')
    sc.build_collated_input_v5(raw_count_data, run_out_dir, collated_path, has_bait2)

    log("  [%s] running DESeq2..." % rid)
    r_output, overdispersion = sc.run_r_script(smg.R_SCRIPT_PATH, collated_path, rscript_exe)
    stats = sc.load_stats_output(os.path.join(run_out_dir, 'everything_combined.csv'))

    log("  [%s] running Bayesian/MCMC (slow step)..." % rid)
    bayes_output, bayes_csv, bayes_overdispersion = sc.run_bayesian_r_script(
        smg.BAYESIAN_R_SCRIPT_PATH, csv_paths, smg.BAYESIAN_THRESHOLD_PPM, run_out_dir, rscript_exe)
    bayesian_stats = sc.load_bayesian_stats_output(bayes_csv)

    log("  [%s] loading junction data..." % rid)
    bqp_data = {}
    for k, p in csv_paths.items():
        basename = os.path.basename(p).replace('_summary.csv', '')
        bqp_path = sc.find_bqp_path(BQP_ROOT, basename)
        bqp_data[k] = sc.load_bqp(bqp_path) if bqp_path else {}

    genes = sorted(raw_count_data['Bait1S'].keys())
    ppm_data = {k: sc.load_ppm_summed(p) for k, p in csv_paths.items()}

    # Actual generation date, not a fixed placeholder - computed now (right
    # before writing), so a batch that crosses midnight names each file by
    # the day it was actually produced.
    date_str = datetime.now().strftime('%d%b%Y')
    name = "SM6_%s_%s" % (rid, date_str)
    tmp_xlsx = os.path.join(run_out_dir, name + "_tmp.xlsx")
    out_csv = os.path.join(OUT_DIR, name + ".csv")

    log("  [%s] building report (temp xlsx)..." % rid)
    smg.Stat_Maker_Gui._write_workbook(None, tmp_xlsx, csv_paths, DATASET_LABELS, overdispersion,
                                       bayes_overdispersion, stats, bayesian_stats, bqp_data,
                                       genes, has_bait2, log, ppm_data=ppm_data)

    log("  [%s] converting to %s ..." % (rid, out_csv))
    xlsx_to_csv(tmp_xlsx, out_csv)
    os.remove(tmp_xlsx)

    return out_csv, len(genes)


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    os.makedirs(INTERMEDIATE_ROOT, exist_ok=True)
    rscript_exe = sc.find_rscript()
    log("===== SM6 batch starting: %d datasets, Rscript=%s =====" % (len(CANONICAL), rscript_exe))

    results = []
    for i, fn in enumerate(CANONICAL, 1):
        rid = rab_id(fn)
        t0 = time.time()
        log("=== [%d/%d] %s (%s) starting ===" % (i, len(CANONICAL), rid, fn))
        try:
            out_csv, n_genes = run_one(fn, rscript_exe)
            elapsed = time.time() - t0
            log("=== [%d/%d] %s DONE in %.0fs (%d genes) -> %s ===" % (i, len(CANONICAL), rid, elapsed, n_genes, out_csv))
            results.append((rid, 'OK', elapsed, ''))
        except Exception as e:
            elapsed = time.time() - t0
            err = "%s: %s" % (type(e).__name__, e)
            log("=== [%d/%d] %s FAILED after %.0fs: %s ===" % (i, len(CANONICAL), rid, elapsed, err))
            log(traceback.format_exc())
            results.append((rid, 'FAILED', elapsed, err))

    log("\n===== BATCH SUMMARY =====")
    n_ok = sum(1 for r in results if r[1] == 'OK')
    log("%d/%d succeeded" % (n_ok, len(results)))
    for rid, status, elapsed, err in results:
        log("  %-10s %-7s %5.0fs  %s" % (rid, status, elapsed, err))
    log("===== SM6 batch complete =====")


if __name__ == '__main__':
    main()
